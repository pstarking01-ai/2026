"""Unit tests for apt_main.py – ApartmentApp."""
import os
import tempfile
import pytest
import pandas as pd

from apt_main import ApartmentApp


@pytest.fixture
def app(qapp):
    """Create a fresh ApartmentApp instance for each test."""
    return ApartmentApp()


# ── Config / type_map ──────────────────────────────────────────────

class TestTypeMap:
    def test_type_map_contains_apt(self, app):
        assert "아파트" in app.type_map
        assert app.type_map["아파트"] == "APT"


# ── auto_hyphen (3-4-4 phone format) ──────────────────────────────

class TestAutoHyphen:
    @pytest.mark.parametrize("raw, expected", [
        ("010", "010"),
        ("01012", "010-12"),
        ("0101234", "010-1234"),
        ("01012345678", "010-1234-5678"),
    ])
    def test_phone_formatting(self, app, raw, expected):
        from PyQt6.QtWidgets import QLineEdit
        w = QLineEdit()
        w.setText(raw)
        app.auto_hyphen(w)
        assert w.text() == expected

    def test_already_formatted(self, app):
        from PyQt6.QtWidgets import QLineEdit
        w = QLineEdit()
        w.setText("010-1234-5678")
        app.auto_hyphen(w)
        assert w.text() == "010-1234-5678"


# ── auto_hyphen_334 (3-3-4 phone format) ─────────────────────────

class TestAutoHyphen334:
    @pytest.mark.parametrize("raw, expected", [
        ("02", "02"),
        ("021", "021"),
        ("0212", "021-2"),
        ("021234", "021-234"),
        ("0212345678", "021-234-5678"),
    ])
    def test_334_formatting(self, app, raw, expected):
        from PyQt6.QtWidgets import QLineEdit
        w = QLineEdit()
        w.setText(raw)
        app.auto_hyphen_334(w)
        assert w.text() == expected


# ── auto_hyphen_date (YYYY-MM-DD) ────────────────────────────────

class TestAutoHyphenDate:
    @pytest.mark.parametrize("raw, expected", [
        ("2025", "2025"),
        ("202506", "2025-06"),
        ("20250622", "2025-06-22"),
        ("2025-06-22", "2025-06-22"),
    ])
    def test_date_formatting(self, app, raw, expected):
        from PyQt6.QtWidgets import QLineEdit
        w = QLineEdit()
        w.setText(raw)
        app.auto_hyphen_date(w)
        assert w.text() == expected

    def test_date_truncates_beyond_8_digits(self, app):
        from PyQt6.QtWidgets import QLineEdit
        w = QLineEdit()
        w.setText("202506221")
        app.auto_hyphen_date(w)
        assert w.text() == "2025-06-22"


# ── format_number (comma formatting) ─────────────────────────────

class TestFormatNumber:
    @pytest.mark.parametrize("raw, expected", [
        ("1000", "1,000"),
        ("50000", "50,000"),
        ("123456789", "123,456,789"),
        ("0", "0"),
    ])
    def test_comma_insertion(self, app, raw, expected):
        from PyQt6.QtWidgets import QLineEdit
        w = QLineEdit()
        w.setText(raw)
        app.format_number(w)
        assert w.text() == expected

    def test_empty_string(self, app):
        from PyQt6.QtWidgets import QLineEdit
        w = QLineEdit()
        w.setText("")
        app.format_number(w)
        assert w.text() == ""

    def test_non_digit_ignored(self, app):
        from PyQt6.QtWidgets import QLineEdit
        w = QLineEdit()
        w.setText("abc")
        app.format_number(w)
        assert w.text() == "abc"

    def test_already_formatted(self, app):
        from PyQt6.QtWidgets import QLineEdit
        w = QLineEdit()
        w.setText("1,000")
        app.format_number(w)
        assert w.text() == "1,000"


# ── calculate_parking ─────────────────────────────────────────────

class TestCalculateParking:
    def test_normal_calculation(self, app):
        app.total_households.setText("100")
        app.total_parking.setText("150")
        app.calculate_parking()
        assert app.parking_per_unit.text() == "1.50"

    def test_fractional_result(self, app):
        app.total_households.setText("300")
        app.total_parking.setText("500")
        app.calculate_parking()
        assert app.parking_per_unit.text() == "1.67"

    def test_invalid_input_clears(self, app):
        app.total_households.setText("")
        app.total_parking.setText("150")
        app.calculate_parking()
        assert app.parking_per_unit.text() == ""


# ── toggle_price_fields ──────────────────────────────────────────

class TestTogglePriceFields:
    def test_buy_mode(self, app):
        app.toggle_price_fields("매매")
        assert not app.price_buy.isHidden()
        assert app.price_jeonse.isHidden()
        assert app.price_deposit.isHidden()
        assert app.price_monthly.isHidden()

    def test_jeonse_mode(self, app):
        app.toggle_price_fields("전세")
        assert app.price_buy.isHidden()
        assert not app.price_jeonse.isHidden()
        assert app.price_deposit.isHidden()
        assert app.price_monthly.isHidden()

    def test_monthly_rent_mode(self, app):
        app.toggle_price_fields("월세")
        assert app.price_buy.isHidden()
        assert app.price_jeonse.isHidden()
        assert not app.price_deposit.isHidden()
        assert not app.price_monthly.isHidden()


# ── refresh_listing_id ────────────────────────────────────────────

class TestRefreshListingId:
    def test_id_format_no_db(self, app, tmp_path):
        app.db_file = str(tmp_path / "nonexistent.xlsx")
        app.refresh_listing_id()
        text = app.listing_id.text()
        assert text.startswith("APT")
        assert "-" in text
        assert text.endswith("-01")

    def test_id_increments_from_db(self, app, tmp_path):
        from datetime import datetime
        date_str = datetime.now().strftime("%y%m%d")
        db_path = str(tmp_path / "estate_db.xlsx")
        existing = pd.DataFrame({"매물ID": [f"APT{date_str}-01", f"APT{date_str}-03"]})
        existing.to_excel(db_path, sheet_name="아파트", index=False)
        app.db_file = db_path
        app.refresh_listing_id()
        assert app.listing_id.text() == f"APT{date_str}-04"


# ── save_sheet ────────────────────────────────────────────────────

class TestSaveSheet:
    def test_creates_new_file(self, app, tmp_path):
        path = str(tmp_path / "test_save.xlsx")
        data = [{"매물ID": "APT-001", "단지명": "테스트"}]
        app.save_sheet(path, "아파트", data)
        assert os.path.exists(path)
        df = pd.read_excel(path, sheet_name="아파트")
        assert len(df) == 1
        assert df.iloc[0]["매물ID"] == "APT-001"

    def test_appends_to_existing_sheet(self, app, tmp_path):
        path = str(tmp_path / "test_save.xlsx")
        data1 = [{"매물ID": "APT-001", "단지명": "A"}]
        app.save_sheet(path, "아파트", data1)
        data2 = [{"매물ID": "APT-002", "단지명": "B"}]
        app.save_sheet(path, "아파트", data2)
        df = pd.read_excel(path, sheet_name="아파트")
        assert len(df) == 2

    def test_updates_existing_row_by_key(self, app, tmp_path):
        path = str(tmp_path / "test_save.xlsx")
        data1 = [{"매물ID": "APT-001", "단지명": "A"}]
        app.save_sheet(path, "아파트", data1)
        data2 = [{"매물ID": "APT-001", "단지명": "A-수정"}]
        app.save_sheet(path, "아파트", data2)
        df = pd.read_excel(path, sheet_name="아파트")
        assert len(df) == 1
        assert df.iloc[0]["단지명"] == "A-수정"

    def test_preserves_other_sheets(self, app, tmp_path):
        path = str(tmp_path / "test_save.xlsx")
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            pd.DataFrame({"col": [1]}).to_excel(writer, sheet_name="기존시트", index=False)
        data = [{"매물ID": "APT-001"}]
        app.save_sheet(path, "아파트", data)
        sheets = pd.read_excel(path, sheet_name=None)
        assert "기존시트" in sheets
        assert "아파트" in sheets


# ── _apply_excel_styles ───────────────────────────────────────────

class TestApplyExcelStyles:
    def test_styles_applied(self, app, tmp_path):
        from openpyxl import load_workbook
        path = str(tmp_path / "styled.xlsx")
        data = [{"매물ID": "APT-001", "단지명": "테스트단지"}]
        app.save_sheet(path, "아파트", data)
        wb = load_workbook(path)
        ws = wb["아파트"]
        header_cell = ws.cell(row=1, column=1)
        assert header_cell.font.bold is True
        assert header_cell.font.size == 16
        data_cell = ws.cell(row=2, column=1)
        assert data_cell.font.bold is False

    def test_column_width_auto_sized(self, app, tmp_path):
        from openpyxl import load_workbook
        path = str(tmp_path / "styled.xlsx")
        data = [{"매물ID": "APT-001", "매우긴한글컬럼": "아주긴내용이들어있는데이터"}]
        app.save_sheet(path, "시트", data)
        wb = load_workbook(path)
        ws = wb["시트"]
        widths = {col_letter: ws.column_dimensions[col_letter].width
                  for col_letter in ["A", "B"]}
        assert widths["B"] > widths["A"]


# ── reset_form ────────────────────────────────────────────────────

class TestResetForm:
    def test_resets_fields(self, app, monkeypatch):
        from PyQt6.QtWidgets import QMessageBox
        monkeypatch.setattr(QMessageBox, "information", lambda *a, **kw: None)
        app.complex_name.setText("테스트단지")
        app.dong.setText("101")
        app.ho.setText("1001")
        app.reset_form()
        assert app.complex_name.text() == ""
        assert app.dong.text() == ""
        assert app.ho.text() == ""
