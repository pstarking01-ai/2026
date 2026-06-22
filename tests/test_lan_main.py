"""Unit tests for lan_main.py – Land (토지) management."""
import os
import sys
import pytest
import pandas as pd
from openpyxl import Workbook

from lan_main import Config, EstateMaster_V13_6


@pytest.fixture
def widget(qapp):
    return EstateMaster_V13_6()


# ── Config ────────────────────────────────────────────────────────

class TestConfig:
    def test_prefix_map(self):
        assert Config.PREFIX_MAP == {"토지": "LND"}

    def test_db_path(self):
        assert Config.DB_PATH.endswith("estate_db.xlsx")

    def test_excel_style(self):
        assert Config.EXCEL_STYLE["font_size"] == 16
        assert Config.EXCEL_STYLE["row_height"] == 30
        assert Config.EXCEL_STYLE["public_color"] == "D9EAD3"
        assert Config.EXCEL_STYLE["private_color"] == "FCE4EC"

    def test_width_constants(self):
        assert Config.WIDTH_KOR == 2.5
        assert Config.WIDTH_ENG == 1.2
        assert Config.MAX_WIDTH == 50

    def test_sns_footer(self):
        assert "성균관" in Config.SNS_FOOTER
        assert "https://pstarking01-ai.github.io/2026/" in Config.SNS_FOOTER


# ── category_config ───────────────────────────────────────────────

class TestCategoryConfig:
    def test_land_category_exists(self, widget):
        assert "토지" in widget.category_config

    def test_land_g2_fields(self, widget):
        g2 = widget.category_config["토지"]["G2"]
        assert "용도지역" in g2
        assert "지목" in g2
        assert "면적(㎡)" in g2

    def test_land_g3_fields(self, widget):
        g3 = widget.category_config["토지"]["G3"]
        assert "진입로(폭)" in g3
        assert "전기/통신" in g3


# ── _find_private_col ─────────────────────────────────────────────

class TestFindPrivateCol:
    def test_basic_detection(self, widget):
        df = pd.DataFrame(columns=["매물ID", "지번", "소유자", "미디어경로(사진/영상)", "면적"])
        prv, med, nums = widget._find_private_col(df)
        assert prv == 2  # "지번" at 1-based index 2
        assert med == 4
        assert 5 in nums  # "면적"

    def test_with_price_columns(self, widget):
        df = pd.DataFrame(columns=["개별공시지가", "매매가", "조정가"])
        _, _, nums = widget._find_private_col(df)
        assert 1 in nums
        assert 2 in nums
        assert 3 in nums

    def test_no_special_columns(self, widget):
        df = pd.DataFrame(columns=["A", "B", "C"])
        prv, med, nums = widget._find_private_col(df)
        assert prv == 999
        assert med == -1
        assert nums == []


# ── _apply_cell_special ───────────────────────────────────────────

class TestApplyCellSpecial:
    def test_hyperlink_for_media(self, widget):
        wb = Workbook()
        ws = wb.active
        ws.append(["http://example.com/photo.jpg"])
        cell = ws.cell(row=1, column=1)
        widget._apply_cell_special(cell, col_idx=1, media_col=1, numeric_cols=[])
        assert cell.hyperlink is not None
        assert cell.font.color.rgb == "000000FF"

    def test_no_hyperlink_for_dash(self, widget):
        wb = Workbook()
        ws = wb.active
        ws.append(["-"])
        cell = ws.cell(row=1, column=1)
        widget._apply_cell_special(cell, col_idx=1, media_col=1, numeric_cols=[])
        assert cell.hyperlink is None

    def test_numeric_format(self, widget):
        wb = Workbook()
        ws = wb.active
        ws.append(["1,234"])
        cell = ws.cell(row=1, column=1)
        widget._apply_cell_special(cell, col_idx=1, media_col=-1, numeric_cols=[1])
        assert cell.value == 1234.0
        assert cell.number_format == '#,##0'

    def test_non_media_non_numeric_unchanged(self, widget):
        wb = Workbook()
        ws = wb.active
        ws.append(["text"])
        cell = ws.cell(row=1, column=1)
        widget._apply_cell_special(cell, col_idx=1, media_col=5, numeric_cols=[3])
        assert cell.value == "text"


# ── _auto_column_width ────────────────────────────────────────────

class TestAutoColumnWidth:
    def test_korean_text_wider(self, widget):
        wb = Workbook()
        ws = wb.active
        ws.append(["ABC", "가나다"])
        widget._auto_column_width(ws)
        w_a = ws.column_dimensions["A"].width
        w_b = ws.column_dimensions["B"].width
        assert w_b > w_a  # Korean chars are wider

    def test_max_width_cap(self, widget):
        wb = Workbook()
        ws = wb.active
        ws.append(["가" * 100])
        widget._auto_column_width(ws)
        assert ws.column_dimensions["A"].width == Config.MAX_WIDTH

    def test_empty_cell_skipped(self, widget):
        wb = Workbook()
        ws = wb.active
        ws.append([None, "A"])
        widget._auto_column_width(ws)
        w_a = ws.column_dimensions["A"].width
        w_b = ws.column_dimensions["B"].width
        assert w_b > 0


# ── _apply_excel_styles ──────────────────────────────────────────

class TestApplyExcelStyles:
    def test_header_bold_data_normal(self, widget):
        wb = Workbook()
        ws = wb.active
        ws.append(["매물ID", "면적"])
        ws.append(["LND-001", "500"])
        df = pd.DataFrame(columns=["매물ID", "면적"])
        widget._apply_excel_styles(ws, df)
        assert ws.cell(1, 1).font.bold is True
        assert ws.cell(2, 1).font.bold is False

    def test_row_height(self, widget):
        wb = Workbook()
        ws = wb.active
        ws.append(["A"])
        ws.append(["B"])
        df = pd.DataFrame(columns=["A"])
        widget._apply_excel_styles(ws, df)
        assert ws.row_dimensions[1].height == Config.EXCEL_STYLE["row_height"]

    def test_public_private_fill(self, widget):
        wb = Workbook()
        ws = wb.active
        ws.append(["매물ID", "지번"])
        ws.append(["LND-001", "123-4"])
        df = pd.DataFrame(columns=["매물ID", "지번"])
        widget._apply_excel_styles(ws, df)
        # "지번" triggers private fill from column 2
        public_fill = ws.cell(2, 1).fill.start_color.rgb
        private_fill = ws.cell(2, 2).fill.start_color.rgb
        assert public_fill != private_fill


# ── format_comma ──────────────────────────────────────────────────

class TestFormatComma:
    def test_formats_number(self, widget):
        from PyQt6.QtWidgets import QLineEdit
        w = QLineEdit()
        widget.all_fields["G2"]["면적(㎡)"] = w
        w.setText("12345")
        widget.format_comma("G2", "면적(㎡)")
        assert w.text() == "12,345"


# ── format_phone ──────────────────────────────────────────────────

class TestFormatPhone:
    @pytest.mark.parametrize("raw, expected", [
        ("01012345678", "010-1234-5678"),
        ("0311234567", "031-1234-567"),
        ("12345", "12345"),
    ])
    def test_phone_formatting(self, widget, raw, expected):
        from PyQt6.QtWidgets import QLineEdit
        w = QLineEdit()
        widget.all_fields["G3"]["연락처"] = w
        w.setText(raw)
        widget.format_phone("G3", "연락처")
        assert w.text() == expected


# ── calculate_land_totals ─────────────────────────────────────────

class TestCalculateLandTotals:
    def test_totals(self, widget):
        from PyQt6.QtWidgets import QTableWidget, QTableWidgetItem, QLineEdit, QLabel
        table = QTableWidget(2, 10)
        # col 7 = area, col 9 = price (per calculate_land_totals)
        table.setItem(0, 7, QTableWidgetItem("500"))
        table.setItem(0, 9, QTableWidgetItem("10000"))
        table.setItem(1, 7, QTableWidgetItem("300"))
        table.setItem(1, 9, QTableWidgetItem("8000"))
        widget.table_land = table

        area_field = QLineEdit()
        price_field = QLineEdit()
        widget.all_fields["G2"]["면적(㎡)"] = area_field
        widget.all_fields["G3"]["매매가(만원)"] = price_field
        widget.lbl_land_totals = QLabel()

        widget.calculate_land_totals()
        text = widget.lbl_land_totals.text()
        assert "800" in text
        assert "18,000" in text
