import sys
import os
import pandas as pd
import datetime
import re
import shutil
from PyQt6.QtWidgets import (QApplication, QWidget, QVBoxLayout, QHBoxLayout, 
                             QLabel, QLineEdit, QPushButton, QTextEdit, QMessageBox, 
                             QGroupBox, QGridLayout, QComboBox, QTableWidget, QTableWidgetItem, QHeaderView,
                             QScrollArea, QFileDialog, QCheckBox, QTabWidget)
from PyQt6.QtCore import Qt
from openpyxl.styles import Font, PatternFill, Alignment

# 시스템 환경 설정 및 상수 관리
class Config:
    DB_PATH = os.path.join(os.path.dirname(__file__), "estate_db.xlsx")

    PHOTO_BASE_DIR = os.path.join(os.path.dirname(__file__), "..", "perporty_photos")
    
    PREFIX_MAP = {
        "토지": "LND"
    }
    SNS_FOOTER = "성균관 공인중개사사무소\n성균관전자명함안내 https://pstarking01-ai.github.io/2026/"
    
    EXCEL_STYLE = {
        "font_size": 16,
        "row_height": 30,
        "public_color": "D9EAD3",
        "private_color": "FCE4EC"
    }
    WIDTH_KOR = 2.5
    WIDTH_ENG = 1.2
    MAX_WIDTH = 50

class EstateMaster_V13_6(QWidget):
    def __init__(self):
        super().__init__()
        self.db_path = Config.DB_PATH
        self.photo_base_dir = Config.PHOTO_BASE_DIR
        self.selected_media = [] # 선택된 사진/영상 경로 리스트

        # 데이터 저장을 위한 필드 딕셔너리 초기화
        self.all_fields = {"G2": {}, "G3": {}, "G4": {}}
        
        # 토지 전용 필드 구성
        self.category_config = {
            "토지": {
                "G2": ["용도지역", "지목","면적(㎡)","개별공시지가(㎡)","매매가(만원))"],
                "G3": ["진입로(폭)", "토목공사", "상/하수도", "전기/통신"]
            }
        }
        self.initUI()

    def initUI(self):
        self.setWindowTitle('성균관 토지 매물 등록 시스템 v1.0')        
        self.setGeometry(100, 100, 1450, 900) # 가로 너비를 조금 더 확장
        self.setStyleSheet("background-color: #f8f9fa; font: 12pt 'Malgun Gothic';")
        
        main_layout = QVBoxLayout()
        

        prop_layout = QVBoxLayout()
        prop_layout.setSpacing(15)

        top_group = QGroupBox("매물 분류 및 위치")
        top_group.setStyleSheet("""
            QGroupBox { border: 2px solid #2c3e50; border-radius: 10px; margin-top: 15px; font-weight: bold; background: white; }
            QGroupBox::title { subcontrol-origin: margin; left: 10px; padding: 0 5px; }
        """)
        v_sel_main = QVBoxLayout()
        v_sel_main.setContentsMargins(15, 25, 15, 15)

        # [1행]: 매물ID 및 조회
        h_sel_1 = QHBoxLayout()
        self.le_m_id = QLineEdit(); self.le_m_id.setPlaceholderText("LND-YYMMDD-NN"); self.le_m_id.setFixedWidth(150)
        self.btn_load = QPushButton("매물 불러오기"); self.btn_load.setFixedWidth(120)
        self.btn_load.clicked.connect(self.load_data)
        self.btn_load.setStyleSheet("background-color: #3498db; color: white; font-weight: bold;")
        h_sel_1.addWidget(QLabel("매물 ID:")); h_sel_1.addWidget(self.le_m_id); h_sel_1.addWidget(self.btn_load)
        h_sel_1.addStretch()
        v_sel_main.addLayout(h_sel_1)

        # [2행]: 주소/위치 입력
        h_sel_2 = QHBoxLayout()
        self.le_do = QLineEdit(); self.le_do.setPlaceholderText("특별/특/광/도"); self.le_do.setFixedWidth(150)
        self.le_si_gun = QLineEdit(); self.le_si_gun.setPlaceholderText("시/군/구"); self.le_si_gun.setFixedWidth(140)
        self.le_emd_ri = QLineEdit(); self.le_emd_ri.setPlaceholderText("읍/면/동/리"); self.le_emd_ri.setFixedWidth(180)
        self.le_jibon = QLineEdit(); self.le_jibon.setPlaceholderText("지번"); self.le_jibon.setFixedWidth(100)
        self.chk_loc_public = QCheckBox("상세정보 공개"); self.chk_loc_public.setChecked(False)

        self.addr_fields = [self.le_do, self.le_si_gun, self.le_emd_ri, self.le_jibon]
        h_sel_2.addWidget(QLabel("주소(위치):"))
        for w in self.addr_fields: h_sel_2.addWidget(w)
        h_sel_2.addWidget(self.chk_loc_public)
        v_sel_main.addLayout(h_sel_2)

        top_group.setLayout(v_sel_main)
        prop_layout.addWidget(top_group)

        self.entry_panel = QHBoxLayout()
        prop_layout.addLayout(self.entry_panel)
        
        self.land_detail_group = self.create_land_detail_ui()
        prop_layout.addWidget(self.land_detail_group)

        # 하단 제어 버튼
        control_h_layout = QHBoxLayout()
        self.btn_copy_markdown = QPushButton("📋 마크다운 테이블 복사")
        self.btn_copy_markdown.setFixedHeight(40)
        self.btn_copy_markdown.setStyleSheet("background-color: #9b59b6; color: white; font-weight: bold;")
        self.btn_copy_markdown.clicked.connect(self.copy_as_markdown)
        control_h_layout.addWidget(self.btn_copy_markdown)
        control_h_layout.addStretch()
        prop_layout.addLayout(control_h_layout)

        self.btn_save = QPushButton("매물등록")
        self.btn_save.clicked.connect(self.process_save)

        # 하단 버튼 및 위젯 멋있게 꾸미기
        self.btn_save.setFixedHeight(50)
        self.btn_save.setStyleSheet("background-color: #2c3e50; color: white; font-weight: bold; font-size: 16px; border-radius: 5px;")

        # 하단 출력창 설정 (목록 확인을 위해 높이 확장)
        self.output = QTextEdit()
        self.output.setFixedHeight(180)
        self.output.setStyleSheet("border: 1px solid #ced4da; border-radius: 5px; background: white; font-size: 13px;")

        btn_h_layout = QHBoxLayout()
        btn_h_layout.addWidget(self.btn_save)

        prop_layout.addLayout(btn_h_layout)
        prop_layout.addWidget(self.output)

        # [스크롤 영역 추가] - 입력 내용이 길어질 경우 버튼이 사라지는 문제 해결
        prop_scroll = QScrollArea()
        prop_container = QWidget()
        prop_container.setLayout(prop_layout)
        prop_scroll.setWidget(prop_container)
        prop_scroll.setWidgetResizable(True)
        prop_scroll.setStyleSheet("border: none;")

        main_layout.addWidget(prop_scroll)

        self.on_type_changed()
        self.setLayout(main_layout)

    def copy_as_markdown(self):
        """현재 토지 매물 정보를 마크다운 테이블 형식으로 클립보드에 복사"""
        try:
            m_id = self.le_m_id.text() or "ID미발급"
            addr = f"{self.le_do.text()} {self.le_si_gun.text()} {self.le_emd_ri.text()} {self.le_jibon.text()}"
            
            md = f"### 📍 토지 매물 상세 보고서 ({m_id})\n\n"
            md += "#### **[ 1. 기본 매물 정보 ]**\n"
            md += "| **구분 항목** | **상세 정보** |\n"
            md += "| :--- | :--- |\n"
            md += f"| **📍 소재지** | **{addr}** |\n"
            
            for group in ["G2", "G3"]:
                for k, v in self.all_fields[group].items():
                    val = v.currentText() if isinstance(v, QComboBox) else v.text()
                    if val and val != "-":
                        md += f"| **{k}** | {val} |\n"
            
            # 필지 상세 정보 테이블 추가
            if self.table_land.rowCount() > 0:
                md += "\n#### **[ 2. 필지별 세부 현황 ]**\n"
                headers = ["**필지ID**", "**지번**", "**용도지역**", "**지목**", "**면적(㎡)**", "**매매가(만원)**", "**소유자**", "**메모**", "**비고**"]
                md += "| " + " | ".join(headers) + " |\n" + "| " + " | ".join(["---"] * len(headers)) + " |\n"
                
                for r in range(self.table_land.rowCount()):
                    row = []
                    # 0:필지ID, 4:지번, 5:용도지역, 6:지목, 7:면적, 9:매매가, 10:소유자, 11:메모, 12:비고
                    for c_idx in [0, 4, 5, 6, 7, 9, 10, 11, 12]:
                        item = self.table_land.item(r, c_idx)
                        row.append(item.text() if item else "")
                    md += "| " + " | ".join(row) + " |\n"

            QApplication.clipboard().setText(md)
            QMessageBox.information(self, "복사 완료", "토지 매물 정보가 마크다운 형식으로 복사되었습니다.")
        except Exception as e:
            QMessageBox.warning(self, "오류", f"마크다운 생성 중 오류: {e}")

    def on_land_table_item_changed(self, item):
        """토지 테이블 수정 시 숫자 포맷팅 및 합계 계산"""
        col = item.column()
        # 7: 면적, 8: 개별공시지가, 9: 매매가
        if col in [7, 8, 9]:
            text = item.text().replace(",", "").replace(" ", "").strip()
            if text.isdigit():
                formatted = format(int(text), ",")
                if item.text() != formatted:
                    self.table_land.blockSignals(True)
                    item.setText(formatted)
                    self.table_land.blockSignals(False)
        self.calculate_land_totals()

    def calculate_land_totals(self):
        """토지 테이블의 총 면적과 총 매매가 합계를 계산"""
        total_area = 0
        total_price = 0
        for r in range(self.table_land.rowCount()):
            try:
                area_item = self.table_land.item(r, 7)
                price_item = self.table_land.item(r, 9)
                if area_item and area_item.text().replace(',', '').strip().isdigit():
                    total_area += int(area_item.text().replace(',', '').strip())
                if price_item and price_item.text().replace(',', '').strip().isdigit():
                    total_price += int(price_item.text().replace(',', '').strip())
            except: continue
            
        # 상단 메인 필드와 자동 동기화
        area_field = self.all_fields.get("G2", {}).get("면적(㎡)")
        price_field = self.all_fields.get("G3", {}).get("매매가(만원)")
        
        if area_field and total_area > 0:
            area_field.blockSignals(True)
            area_field.setText(format(total_area, ","))
            area_field.blockSignals(False)
        if price_field and total_price > 0:
            price_field.blockSignals(True)
            price_field.setText(format(total_price, ","))
            price_field.blockSignals(False)
            
        self.lbl_land_totals.setText(f"합계: 총 면적 {total_area:,}㎡ / 총 매매가 {total_price:,}만원")

    def add_land_row(self):
        self.table_land.insertRow(self.table_land.rowCount())

    def del_land_row(self):
        curr = self.table_land.currentRow()
        if curr >= 0: self.table_land.removeRow(curr)

    def clear_inputs(self):
        """매물 등록 후 모든 입력 필드 초기화"""
        self.le_m_id.clear()
        self.le_do.clear()
        self.le_si_gun.clear()
        self.le_emd_ri.clear()
        self.le_jibon.clear()
        for group in self.all_fields.values():
            for widget in group.values():
                if isinstance(widget, QLineEdit): widget.clear()
                elif isinstance(widget, QComboBox): widget.setCurrentIndex(0)
        self.table_land.setRowCount(0)
        self.calculate_land_totals()

    def init_activity_tab(self):
        """임장활동 UI 초기화 (필요시 호출)"""
        prep_group = QGroupBox("📋 임장 준비")
        prep_group.setStyleSheet("QGroupBox { font-weight: bold; border: 1px solid #ced4da; background: #fdfefe; margin-top: 10px; }")
        prep_layout = QGridLayout()
        prep_layout.setContentsMargins(15, 25, 15, 15)
        prep_layout.setVerticalSpacing(10)
        self.te_recommend_list = QTextEdit(); self.te_recommend_list.setPlaceholderText("중개사 추천매물 List up (매물 ID 또는 간략 정보)"); self.te_recommend_list.setFixedHeight(80)
        self.le_view_pw = QLineEdit(); self.le_view_pw.setPlaceholderText("비밀번호 (공실)")
        self.chk_view_confirmed = QCheckBox("집보기 예약 확인")
        self.le_owner_contact_act = QLineEdit(); self.le_owner_contact_act.setPlaceholderText("소유자 연락처 확인")

    def select_media(self):
        """파일 다이얼로그를 통해 사진 및 영상 선택"""
        files, _ = QFileDialog.getOpenFileNames(self, "매물 사진/영상 선택", "", "Media Files (*.png *.jpg *.jpeg *.gif *.mp4 *.mov *.avi *.mkv)")
        if files:
            self.selected_media = files

    def create_land_detail_ui(self):
        """토지 상세 내역(여러 필지) 입력을 위한 UI 그룹 생성"""
        group = QGroupBox("토지접수 상세 (필지별 현황/단위:㎡/만원)")
        group.setStyleSheet("QGroupBox { font-weight: bold; border: 1px solid #ced4da; margin-top: 30px; }")
        layout = QVBoxLayout()
        layout.setContentsMargins(15, 35, 15, 15)
        layout.setSpacing(10)

        self.table_land = QTableWidget(0, 13)
        self.table_land.setHorizontalHeaderLabels(["필지ID", "특별/특/광/도", "시/군/구", "읍/면/동/리", "지번", "용도지역", "지목", "면적(㎡)", "개별공시지가(㎡,원)", "매매가(만원)", "소유자", "메모", "비고"])
        self.table_land.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.table_land.horizontalHeader().setStretchLastSection(True)
        self.table_land.setFixedHeight(200)
        self.table_land.itemChanged.connect(self.on_land_table_item_changed)

        btn_layout = QHBoxLayout()
        self.btn_add_row_l = QPushButton("+ 필지 추가")
        self.btn_del_row_l = QPushButton("- 필지 삭제")
        self.btn_add_row_l.clicked.connect(self.add_land_row)
        self.btn_del_row_l.clicked.connect(self.del_land_row)
        btn_layout.addWidget(self.btn_add_row_l)
        btn_layout.addWidget(self.btn_del_row_l)
        btn_layout.addStretch()
        
        self.lbl_land_totals = QLabel("합계: 총 면적 0㎡ / 총 매매가 0만원")
        self.lbl_land_totals.setStyleSheet("font-weight: bold; color: #27ae60; font-size: 14px;")
        
        layout.addWidget(self.table_land)
        layout.addLayout(btn_layout)
        layout.addWidget(self.lbl_land_totals)
        group.setLayout(layout)
        group.setVisible(False)
        return group
    def generate_marketing_text(self):
        """공개 설정 여부를 반영한 마케팅용 발송 문구 생성"""
        type_name = "토지"
        m_id = self.le_m_id.text() or "ID 미발급"
        is_public = self.chk_loc_public.isChecked()
        
        addr = f"{self.le_do.text()} {self.le_si_gun.text()} {self.le_emd_ri.text()}"
        jibon = self.le_jibon.text()
        detail_loc = f"\n📍 위치: {addr} " + (f"{jibon}" if is_public else "(상세지번 비공개)")

        # 주요 정보 요약 (G2, G3에서 발췌)
        summary = []
        for group in ["G2", "G3"]:
            for k, v in self.all_fields[group].items():
                val = v.currentText() if isinstance(v, QComboBox) else v.text()
                if val and val != "-":
                    summary.append(f"▪ {k}: {val}")

        text = f"""[성균관 부동산 매물안내]
━━━━━━━━━━━━━━
🏠 매물종류: {type_name}
🆔 매물번호: {m_id}{detail_loc}

📋 주요정보
{chr(10).join(summary)}

━━━━━━━━━━━━━━
✨ {type_name} 전문 중개! 
궁금하신 사항은 언제든 문의주세요.\n{Config.SNS_FOOTER}"""
        self.output.clear()
        self.output.append(text)
        QMessageBox.information(self, "문구 생성 완료", "하단 출력창에 발송용 문구가 생성되었습니다. 복사해서 사용하세요.")

    def add_to_multi_list(self):
        """현재 입력된 매물 정보를 요약하여 목록형 문구로 누적 추가"""
        type_name = "토지"
        m_id = self.le_m_id.text().strip().upper() or "ID미발급"
        addr = self.le_emd_ri.text() or "주소미입력"
        
        # 주요 수치 추출 (면적 및 금액)
        area_w = self.all_fields["G2"].get("면적(㎡)") or self.all_fields["G2"].get("대지면적(㎡)")
        price_w = self.all_fields["G3"].get("매매가(만원)") or self.all_fields["G3"].get("매매가") or self.all_fields["G3"].get("희망매매가(만원)")
        
        area = area_w.text() if area_w else "-"
        price = price_w.text() if price_w else "-"
        
        # 토지의 경우 지목 추가
        extra = ""
        if type_name == "토지":
            jm_w = self.all_fields["G2"].get("지목")
            jm = (jm_w.currentText() if isinstance(jm_w, QComboBox) else jm_w.text()) if jm_w else ""
            if jm and jm != "-": extra = f" / {jm}"

        current_text = self.output.toPlainText()
        header = "📢 [성균관 부동산 추천 매물 리스트]\n━━━━━━━━━━━━━━\n"
        footer = f"\n━━━━━━━━━━━━━━\n✨ 문의주시면 친절히 안내해 드립니다.{Config.SNS_FOOTER}"
        
        # 이미 목록이 작성 중인지 확인
        if "추천 매물 리스트" not in current_text:
            new_content = header
        else:
            # 기존 하단 안내 문구(footer) 제거 후 내용만 추출
            new_content = current_text.split("━━━━━━━━━━━━━━\n✨")[0]
            
        item_line = f"📍 {type_name}({m_id}): {addr} | {area}㎡ | {price}만원{extra}\n"
        
        self.output.clear()
        self.output.setText(new_content + item_line + footer)
        self.tabs.setCurrentIndex(0)
        QMessageBox.information(self, "목록 추가", f"{m_id} 매물이 안내 목록에 추가되었습니다.")

    def load_data(self):
        """매물ID를 기반으로 엑셀에서 데이터를 찾아 입력창에 채움"""
        m_id = self.le_m_id.text().strip().upper()
        if not m_id:
            QMessageBox.warning(self, "알림", "조회할 매물ID를 입력해주세요.")
            return

        if not os.path.exists(self.db_path):
            QMessageBox.warning(self, "오류", "데이터베이스 파일이 존재하지 않습니다.")
            return

        try:
            found = False
            with pd.ExcelFile(self.db_path, engine='openpyxl') as reader:
                # 1. 메인 정보 조회 (토지 고정)
                if "토지" in reader.sheet_names:
                    df = pd.read_excel(reader, sheet_name="토지")
                    row = df[df['매물ID'] == m_id]

                    if not row.empty:
                        found = True
                        data = row.iloc[0].to_dict()
                        self.le_do.setText(str(data.get("특별/특/광/도", "")))
                        self.le_si_gun.setText(str(data.get("시/군/구", "")))
                        self.le_emd_ri.setText(str(data.get("읍/면/동/리", "")))
                        self.le_jibon.setText(str(data.get("지번", "")))
                        
                        for group_name in ["G2", "G3", "G4"]:
                            for label, widget in self.all_fields[group_name].items():
                                val = str(data.get(label, ""))
                                if val == "nan": val = ""
                                if isinstance(widget, QComboBox): widget.setCurrentText(val)
                                else: widget.setText(val)

                # 2. 토지 상세 테이블 조회 (토지상세 고정)
                if "토지상세" in reader.sheet_names:
                    df_detail = pd.read_excel(reader, sheet_name="토지상세")
                    rows = df_detail[df_detail['매물ID'] == m_id]
                    self.table_land.setRowCount(0)
                    headers = ["필지ID", "특별/특/광/도", "시/군/구", "읍/면/동/리", "지번", "용도지역", "지목", "면적(㎡)", "개별공시지가(㎡,원)", "매매가(만원)", "소유자", "메모", "비고"]
                    for _, d_row in rows.iterrows():
                        r_idx = self.table_land.rowCount()
                        self.table_land.insertRow(r_idx)
                        for c_idx, h in enumerate(headers):
                            val = str(d_row.get(h, ""))
                            if val == "nan": val = ""
                            self.table_land.setItem(r_idx, c_idx, QTableWidgetItem(val))

            if found:
                self.output.append(f"✅ {m_id} 데이터를 성공적으로 불러왔습니다.")
            else:
                QMessageBox.warning(self, "실패", f"ID '{m_id}'에 해당하는 매물을 찾을 수 없습니다.")
        except Exception as e:
            QMessageBox.critical(self, "오류", f"조회 중 오류 발생: {str(e)}")

    def create_box(self, title, color, labels):
        """글자 겹침 방지를 위해 간격과 여백을 재설계한 함수"""
        box = QGroupBox(title)
        # 1. GroupBox 스타일: margin-top을 40px로 늘려 제목 겹침 방지
        box.setStyleSheet(f"""
            QGroupBox {{ 
                background-color: {color}; 
                border: 1px solid #c0c0c0; 
                border-radius: 8px; 
                margin-top: 20px; 
                font-weight: bold; 
            }}
            QGroupBox::title {{ 
                subcontrol-origin: margin; 
                subcontrol-position: top left;
                left: 10px; 
                padding: 0 10px; 
                top: 0px; 
                font-size: 14px;
                color: #333333;
            }}
        """)
        
        layout = QGridLayout()
        # 행 간격 및 내부 여백 최적화
        layout.setVerticalSpacing(15) # 간격을 12에서 15로 조정하여 가독성 향상
        layout.setHorizontalSpacing(10)
        layout.setContentsMargins(15, 30, 15, 15)
        
        # 드롭다운 데이터 설정
        dropdown_data = {
            
            "광고상황": ["광고전", "진행중", "거래완료", "매물철회"],
            "광고사이트": ["한방", "이실장", "부동산써브"],
            "자체광고 채널": ["홈페이지", "블로그", "유튜브"],
            "용도지역": ["제1종전용주거", "제2종전용주거", "제1종일반주거", "제2종일반주거", "제3종일반주거", "준주거", "중심상업", "일반상업", "근린상업", "유통상업", "전용공업지역", "일반공업지역", "준공업지역","보전관리지역", "생산관리지역", "계획관리지역", "농림지역", "자연환경보전지역"],
            "통신사": ["SKT", "KT", "LGU+", "SKT알뜰폰", "KT알뜰폰", "LGU+알뜰폰", "미확인"],
            "연락처구분": ["소유자", "관계인", "기타"]
        }

        field_dict = {}
        for i, text in enumerate(labels):
            lbl = QLabel(text)
            lbl.setMinimumWidth(90)
            # 라벨이 입력창 중앙에 오도록 수직 정렬
            lbl.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            
            if text == "사진,영상링크":
                edit = QPushButton(f"사진,영상링크 ({len(self.selected_media)})")
                edit.setFixedHeight(32)
                edit.setStyleSheet("background-color: #f39c12; color: white; font-weight: bold; border-radius: 4px;")
                edit.clicked.connect(self.select_media)
                self.btn_add_media = edit
            elif any(key in text for key in dropdown_data):
                # 부분 일치하는 키를 찾아 드롭다운 적용 (예: "용도지역" -> "용도지역")
                match_key = next(key for key in dropdown_data if key in text)
                edit = QComboBox()
                edit.addItems(dropdown_data[match_key])
                edit.setFixedHeight(32)
                edit.setStyleSheet("background-color: white; border: 1px solid #ced4da; border-radius: 4px;")
            else:
                edit = QLineEdit()
                edit.setFixedHeight(32)
                edit.setStyleSheet("background-color: white; border: 1px solid #ced4da; border-radius: 4px; padding: 2px 5px;")
            
            field_dict[text] = edit
            layout.addWidget(lbl, i, 0)
            layout.addWidget(edit, i, 1)
            
        box.setLayout(layout)
        return box, field_dict

    def on_type_changed(self):
        """매물 종류 변경 시 입력 필드를 동적으로 재생성"""
        # 기존 위젯 및 레이아웃 제거
        while self.entry_panel.count():
            item = self.entry_panel.takeAt(0)
            if item.layout():
                while item.layout().count():
                    sub_item = item.layout().takeAt(0)
                    if sub_item.widget():
                        sub_item.widget().deleteLater()
            if item.widget(): item.widget().deleteLater()

        type_name = "토지"
        self.land_detail_group.setVisible(True)

        cfg = self.category_config[type_name]
        self.box2, self.all_fields["G2"] = self.create_box("상세정보", "#f1f8e9", cfg["G2"])
        self.box3, self.all_fields["G3"] = self.create_box("기반시설", "#fff3e0", cfg["G3"])

        g4_labels = ["소유자", "성별", "소유자연락처", "관계인연락처", "조정가", "중개사메모1", "중개사메모2"]
        self.box4, self.all_fields["G4"] = self.create_box("미공개정보", "#fce4ec", g4_labels)

        # 상세정보(G2)와 가격정보(G3)를 한 열(수직)로 배치
        col1_vbox = QVBoxLayout()
        col1_vbox.addWidget(self.box2)
        col1_vbox.addWidget(self.box3)
        
        self.entry_panel.addLayout(col1_vbox)
        self.entry_panel.addWidget(self.box4)

        # 실시간 포맷팅 연결 (콤마 및 하이픈)
        if "매매가(만원)" in self.all_fields["G3"]:
            self.all_fields["G3"]["매매가(만원)"].textChanged.connect(lambda _: self.format_comma("G3", "매매가(만원)"))
            self.all_fields["G3"]["매매가(만원)"].textChanged.connect(self.calculate_land_totals)

        if "면적(㎡)" in self.all_fields["G2"]:
            self.all_fields["G2"]["면적(㎡)"].textChanged.connect(lambda _: self.format_comma("G2", "면적(㎡)"))

        if "개별공시지가(㎡)" in self.all_fields["G3"]:
            self.all_fields["G3"]["개별공시지가(㎡)"].textChanged.connect(lambda _: self.format_comma("G3", "개별공시지가(㎡)"))

        if "조정가" in self.all_fields["G4"]:
            self.all_fields["G4"]["조정가"].textChanged.connect(lambda _: self.format_comma("G4", "조정가"))
            
        if "소유자연락처" in self.all_fields["G4"]:
            self.all_fields["G4"]["소유자연락처"].textChanged.connect(lambda _: self.format_phone("G4", "소유자연락처"))
        if "관계인연락처" in self.all_fields["G4"]:
            self.all_fields["G4"]["관계인연락처"].textChanged.connect(lambda _: self.format_phone("G4", "관계인연락처"))

    def format_comma(self, group, field):
        """숫자 입력 시 실시간으로 콤마(,) 추가"""
        edit = self.all_fields[group][field]
        text = edit.text().replace(",", "")
        if text.isdigit():
            edit.blockSignals(True)
            try: edit.setText(format(int(text), ","))
            except: pass
            edit.blockSignals(False)

    def format_phone(self, group, field_name):
        """전화번호 입력 시 하이픈(-) 자동 삽입"""
        edit = self.all_fields[group][field_name]
        text = re.sub(r'[^0-9]', '', edit.text())
        if len(text) >= 10:
            formatted = f"{text[:3]}-{text[3:7]}-{text[7:]}"
        elif len(text) > 6:
            formatted = f"{text[:3]}-{text[3:6]}-{text[6:]}"
        else: formatted = text
        edit.blockSignals(True)
        edit.setText(formatted)
        edit.blockSignals(False)

    def process_save(self):
        type_name = "토지"
        pfx = "LND"
        m_id = self.le_m_id.text().strip().upper()

        if not m_id:
            # 신규 ID 생성 로직
            date_str = datetime.datetime.now().strftime("%y%m%d")
            seq = 1
            if os.path.exists(self.db_path):
                try:
                    with pd.ExcelFile(self.db_path, engine='openpyxl') as reader:
                        max_seq = 0
                        date_pattern = f"-{date_str}-"
                        for s_name in reader.sheet_names:
                            df_tmp = pd.read_excel(reader, sheet_name=s_name)
                            if '매물ID' in df_tmp.columns:
                                matched_ids = df_tmp[df_tmp['매물ID'].astype(str).str.contains(date_pattern)]['매물ID'].tolist()
                                for mid in matched_ids:
                                    try:
                                        num = int(str(mid).split('-')[-1])
                                        if num > max_seq: max_seq = num
                                    except: continue
                        seq = max_seq + 1
                except: pass
            m_id = f"{pfx}-{date_str}-{seq:02d}"

        media_path_info = "-" # 마케팅 정보 삭제로 인해 고정값 처리

        # 2. 데이터 수집
        record = {"매물ID": m_id}
        is_public = self.chk_loc_public.isChecked()
        record.update({"특별/특/광/도": self.le_do.text(), "시/군/구": self.le_si_gun.text(), "읍/면/동/리": self.le_emd_ri.text()})

        if is_public: record["지번"] = self.le_jibon.text()

        for group_key in ["G2", "G3"]:
            for label, widget in self.all_fields[group_key].items():
                if not isinstance(widget, QPushButton):
                    record[label] = widget.currentText() if isinstance(widget, QComboBox) else widget.text()

        record["미디어경로(사진/영상)"] = media_path_info
        if not is_public: record["지번"] = self.le_jibon.text()

        for label, widget in self.all_fields["G4"].items():
            record[label] = widget.currentText() if isinstance(widget, QComboBox) else widget.text()

        # 3. 상세 내역 추출 (토지상세)
        detail_df_land = None
        if self.table_land.rowCount() > 0:
            land_details = []
            l_headers = ["필지ID", "특별/특/광/도", "시/군/구", "읍/면/동/리", "지번", "용도지역", "지목", "면적(㎡)", "개별공시지가(㎡,원)", "매매가(만원)", "소유자", "메모", "비고"]
            for r in range(self.table_land.rowCount()):
                sub_id = f"{m_id}-{r+1:02d}" # 하위 코드 생성 (LND-260522-01-01)
                row_data = {"매물ID": m_id}
                for c in range(self.table_land.columnCount()):
                    item = self.table_land.item(r, c)
                    val = sub_id if c == 0 else (item.text() if item else "")
                    row_data[l_headers[c]] = val
                land_details.append(row_data)
            detail_df_land = pd.DataFrame(land_details)

        try:
            self.output.append(f"⏳ {m_id} 데이터를 저장 중입니다...")
            QApplication.processEvents() 
            
            # 데이터 수집 순서(입력창 순서)가 반영된 record 기반 DataFrame 생성
            new_df = pd.DataFrame([record])
            sheets = {}

            try:
                if os.path.exists(self.db_path):
                    with pd.ExcelFile(self.db_path, engine='openpyxl') as reader:
                        sheets = {sheet: pd.read_excel(reader, sheet_name=sheet) for sheet in reader.sheet_names}
                    
                    if type_name in sheets:
                        # 기존에 동일 ID가 있으면 삭제 (수정 대응)
                        df_main = sheets[type_name]
                        df_main = df_main[df_main['매물ID'] != m_id]
                        combined = pd.concat([df_main, new_df], ignore_index=True)
                        cols = list(record.keys())
                        for c in combined.columns:
                            if c not in cols: cols.append(c)
                        sheets[type_name] = combined[cols]
                    else:
                        sheets[type_name] = new_df
                    
                    # 토지 상세 DB 처리
                    if detail_df_land is not None and not detail_df_land.empty:
                        l_sheet = "토지상세"
                        if l_sheet in sheets:
                            sheets[l_sheet] = sheets[l_sheet][sheets[l_sheet]['매물ID'] != m_id]
                            sheets[l_sheet] = pd.concat([sheets[l_sheet], detail_df_land], ignore_index=True)
                        else:
                            sheets[l_sheet] = detail_df_land
                else:
                    sheets = {type_name: new_df}
                    if detail_df_land is not None:
                        sheets["토지상세"] = detail_df_land

                # 등록된 시트들 중에서 상담일지 등이 사라지지 않도록 유지
                # (reader에서 가져온 기존 시트들은 이미 sheets에 담겨 있음)

                # 엑셀 저장 및 서식 적용
                with pd.ExcelWriter(self.db_path, engine='openpyxl') as writer:
                    for s_name, df in sheets.items():
                        df.to_excel(writer, sheet_name=s_name, index=False)
                        ws = writer.sheets[s_name]
                        self._apply_excel_styles(ws, df)
            except PermissionError:
                raise Exception(f"엑셀 파일('{self.db_path}')이 이미 열려 있습니다.\n엑셀 프로그램을 닫고 다시 시도해 주세요.")

            loc_info = f"{record['시/군/구']} {record['읍/면/동/리']} {record.get('지번', '')}"
            self.output.append(f"✅ {m_id} 저장 성공: {loc_info}")
            QMessageBox.information(self, "성공", f"{m_id} 저장 및 분석이 완료되었습니다.")
            self.clear_inputs() # 저장 성공 후 입력창 초기화
        except Exception as e:
            self.output.append(f"❌ 에러 발생: {str(e)}")
            QMessageBox.critical(self, "저장 에러", f"데이터를 저장하지 못했습니다.\n원인: {str(e)}")

    def _find_private_col(self, df):
        """비공개 정보 시작점 및 특수 컬럼 인덱스 분석"""
        private_col_idx = 999
        media_col_idx = -1
        numeric_cols = []
        for i, col_name in enumerate(df.columns, 1):
            name_str = str(col_name)
            if name_str in ["지번", "소유자"]: private_col_idx = min(private_col_idx, i)
            if name_str == "미디어경로(사진/영상)": media_col_idx = i
            if any(kw in name_str for kw in ["면적", "개별공시지가", "매매가", "조정가"]):
                numeric_cols.append(i)
        return private_col_idx, media_col_idx, numeric_cols

    def _apply_excel_styles(self, ws, df):
        """엑셀 워크시트에 서식을 적용하는 전용 함수"""
        prv_idx, med_idx, num_cols = self._find_private_col(df)
        style = Config.EXCEL_STYLE
        font_normal = Font(name='맑은 고딕', size=style["font_size"])
        font_bold = Font(name='맑은 고딕', size=style["font_size"], bold=True)
        fill_public = PatternFill(start_color=style["public_color"], end_color=style["public_color"], fill_type="solid")
        fill_private = PatternFill(start_color=style["private_color"], end_color=style["private_color"], fill_type="solid")

        for r_idx, row in enumerate(ws.iter_rows(), 1):
            ws.row_dimensions[r_idx].height = style["row_height"]
            is_header = (r_idx == 1)
            for c_idx, cell in enumerate(row, 1):
                cell.font = font_bold if is_header else font_normal
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.fill = fill_private if c_idx >= prv_idx else fill_public
                
                if not is_header:
                    self._apply_cell_special(cell, c_idx, med_idx, num_cols)
        
        self._auto_column_width(ws)

    def _apply_cell_special(self, cell, col_idx, media_col, numeric_cols):
        """개별 셀 특수 처리 (하이퍼링크 및 숫자 포맷)"""
        if col_idx == media_col and cell.value and cell.value != "-":
            cell.hyperlink = cell.value
            cell.font = Font(name='맑은 고딕', size=Config.EXCEL_STYLE["font_size"], color="0000FF", underline="single")
        elif col_idx in numeric_cols and cell.value and cell.value != "-":
            try:
                cell.value = float(str(cell.value).replace(",", ""))
                cell.number_format = '#,##0'
            except: pass

    def _auto_column_width(self, ws):
        """내용에 따른 열 너비 자동 조정"""
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    current_len = sum(Config.WIDTH_KOR if ord(c) > 128 else Config.WIDTH_ENG for c in str(cell.value))
                    if current_len > max_len: max_len = current_len
            ws.column_dimensions[col_letter].width = min(max_len + 2, Config.MAX_WIDTH)


if __name__ == '__main__':
    app = QApplication(sys.argv); window = EstateMaster_V13_6(); window.show(); sys.exit(app.exec())