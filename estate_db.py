"""
estate_db.py - 부동산 중개업무자동화시스템 공통 모듈
공통 함수 및 설정을 제공합니다.
"""
import os
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment


# 공통 엑셀 스타일 설정
EXCEL_STYLE = {
    "font_size": 16,
    "row_height": 30,
    "public_color": "D9EAD3",
    "private_color": "FCE4EC"
}
WIDTH_KOR = 2.5
WIDTH_ENG = 1.2
MAX_WIDTH = 50


def update_apartment_in_excel(db_path, complex_name, dong, ho, update_data):
    """
    estate_db.xlsx에서 아파트 매물 정보를 업데이트합니다.
    
    Args:
        db_path: 엑셀 DB 파일 경로
        complex_name: 단지명
        dong: 동
        ho: 호
        update_data: 업데이트할 데이터 딕셔너리 (예: {"공급면적": "84", "전용면적": "59"})
    """
    if not os.path.exists(db_path):
        return False
    
    try:
        all_sheets = pd.read_excel(db_path, sheet_name=None, engine='openpyxl')
        sheet_name = "아파트"
        
        if sheet_name not in all_sheets:
            return False
        
        df = all_sheets[sheet_name]
        
        # 단지명, 동, 호가 모두 일치하는 행 찾기
        mask = (
            (df["단지명"].astype(str).str.strip() == str(complex_name).strip()) &
            (df["동"].astype(str).str.strip() == str(dong).strip()) &
            (df["호"].astype(str).str.strip() == str(ho).strip())
        )
        
        if mask.any():
            for col, val in update_data.items():
                if col in df.columns:
                    df.loc[mask, col] = val
            
            all_sheets[sheet_name] = df
            
            with pd.ExcelWriter(db_path, engine='openpyxl') as writer:
                for s_name, s_df in all_sheets.items():
                    s_df.to_excel(writer, sheet_name=s_name, index=False)
            return True
    except Exception as e:
        print(f"DB 업데이트 오류: {e}")
    
    return False


def find_private_col(df):
    """비공개 정보 시작점 및 특수 컬럼 인덱스 분석"""
    private_col_idx = 999
    media_col_idx = -1
    numeric_cols = []
    for i, col_name in enumerate(df.columns, 1):
        name_str = str(col_name)
        if name_str in ["지번", "소유자"]:
            private_col_idx = min(private_col_idx, i)
        if name_str == "미디어경로(사진/영상)":
            media_col_idx = i
        if any(kw in name_str for kw in ["면적", "개별공시지가", "매매가", "보증금", "월세", "조정가"]):
            numeric_cols.append(i)
    return private_col_idx, media_col_idx, numeric_cols


def apply_excel_styles(ws, df):
    """엑셀 워크시트에 서식을 적용하는 공통 함수"""
    prv_idx, med_idx, num_cols = find_private_col(df)
    font_normal = Font(name='맑은 고딕', size=EXCEL_STYLE["font_size"])
    font_bold = Font(name='맑은 고딕', size=EXCEL_STYLE["font_size"], bold=True)
    fill_public = PatternFill(start_color=EXCEL_STYLE["public_color"], end_color=EXCEL_STYLE["public_color"], fill_type="solid")
    fill_private = PatternFill(start_color=EXCEL_STYLE["private_color"], end_color=EXCEL_STYLE["private_color"], fill_type="solid")

    for r_idx, row in enumerate(ws.iter_rows(), 1):
        ws.row_dimensions[r_idx].height = EXCEL_STYLE["row_height"]
        is_header = (r_idx == 1)
        for c_idx, cell in enumerate(row, 1):
            cell.font = font_bold if is_header else font_normal
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.fill = fill_private if c_idx >= prv_idx else fill_public
            
            if not is_header:
                # 하이퍼링크 및 숫자 포맷 처리
                if c_idx == med_idx and cell.value and cell.value != "-":
                    cell.hyperlink = str(cell.value)
                    cell.font = Font(name='맑은 고딕', size=EXCEL_STYLE["font_size"], color="0000FF", underline="single")
                elif c_idx in num_cols and cell.value and cell.value != "-":
                    try:
                        cell.value = float(str(cell.value).replace(",", ""))
                        cell.number_format = '#,##0'
                    except:
                        pass
    
    # 자동 열 너비 조정
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                current_len = sum(WIDTH_KOR if ord(c) > 128 else WIDTH_ENG for c in str(cell.value))
                if current_len > max_len:
                    max_len = current_len
        ws.column_dimensions[col_letter].width = min(max_len + 2, MAX_WIDTH)


def open_file_cross_platform(file_path):
    """크로스 플랫폼 파일 열기"""
    import subprocess
    import sys
    
    if sys.platform == 'win32':
        os.startfile(file_path)
    elif sys.platform == 'darwin':
        subprocess.call(['open', file_path])
    else:
        subprocess.call(['xdg-open', file_path])
