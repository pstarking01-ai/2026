import sys
import os
import pandas as pd
import datetime
import re
import shutil
from reception.estate_db import update_apartment_in_excel
from PyQt6.QtWidgets import (QApplication, QWidget, QVBoxLayout, QHBoxLayout, 
                             QLabel, QLineEdit, QPushButton, QTextEdit, QMessageBox, 
                             QGroupBox, QGridLayout, QComboBox, QTableWidget, QTableWidgetItem, QHeaderView,
                             QScrollArea, QFileDialog, QCheckBox, QTabWidget)
from PyQt6.QtCore import Qt
from openpyxl.styles import Font, PatternFill, Alignment

# 시스템 환경 설정 및 상수 관리
class Config:
    DB_PATH = os.path.join(os.path.dirname(__file__), "estate_db.xlsx")
    PHOTO_BASE_DIR = os.path.join(os.path.dirname(__file__), "property_photos")
    
    PREFIX_MAP = {
        "공장": "FCT"
    }
    SNS_FOOTER = "성균관 공인중개사사무소\n성균관전자명함안내 https://pstarking01-ai.github.io/2026/"

class EstateMaster_V13_6(QWidget):
    def __init__(self):
        super().__init__()
        self.db_path = Config.DB_PATH
        self.photo_base_dir = Config.PHOTO_BASE_DIR
        self.selected_media = [] # 선택된 사진/영상 경로 리스트
        self.all_fields = {"G2": {}, "G3": {}, "MKT": {}, "G4": {}}
        self.category_config = {
            "공장": {
                "G2": ["사용승인일", "대지면적(㎡)", "건축면적(㎡)", "연면적(㎡)", "용도지역/지구/구역", "지목", "주구조", "주용도", "건물높이", "층수"],
                "G3": ["거래가", "보증금", "월세", "화장실", "승강기", "주차장(대)", "호이스트", "도로조건", "하수처리", "전력(kw/h)"],
                "MKT": ["광고상황", "광고사이트", "자체광고 채널", "사진,영상링크"]
            }
        }
        self.initUI()

    def initUI(self):
        self.setWindowTitle('성균관 부동산 Master_system v13.6')        
        self.setGeometry(100, 100, 1450, 900) # 가로 너비를 조금 더 확장
        self.setStyleSheet("background-color: #f8f9fa; font: 12pt 'Malgun Gothic';")
        
        main_layout = QVBoxLayout()
        
        # [탭 위젯 생성]
        self.tabs = QTabWidget()
        self.property_tab = QWidget()
        self.analysis_tab = QWidget()
        self.marketing_tab = QWidget()
        self.consult_tab = QWidget()
        self.activity_tab = QWidget()
        self.contract_tab = QWidget()
        
        self.tabs.addTab(self.property_tab, "🏠 매물 관리")
        self.tabs.addTab(self.analysis_tab, "📈 데이터 분석")
        self.tabs.addTab(self.marketing_tab, "📊 마케팅")
        self.tabs.addTab(self.consult_tab, " 🧑‍💼 고객상담")
        self.tabs.addTab(self.activity_tab, "🏃 임장활동")
        self.tabs.addTab(self.contract_tab, " 📄 계약관리")


        
        # --- 1번 탭: 매물 관리 (기존 레이아웃) ---
        prop_layout = QVBoxLayout()
        prop_layout.setSpacing(15)

        top_group = QGroupBox("매물 분류,위치 (Smart Selector)")
        top_group.setStyleSheet("""
            QGroupBox { border: 2px solid #2c3e50; border-radius: 10px; margin-top: 15px; font-weight: bold; background: white; }
            QGroupBox::title { subcontrol-origin: margin; left: 10px; padding: 0 5px; }
        """)
        v_sel_main = QVBoxLayout()
        v_sel_main.setContentsMargins(15, 25, 15, 15)
        v_sel_main.setSpacing(10)

        # [1행]: 매물ID 및 조회
        h_sel_1 = QHBoxLayout()
        self.le_m_id = QLineEdit(); self.le_m_id.setPlaceholderText("매물ID (수정시 입력)"); self.le_m_id.setFixedWidth(130)
        self.btn_load = QPushButton("조회/수정"); self.btn_load.setFixedWidth(80)
        self.btn_load.clicked.connect(self.load_data)
        self.btn_load.setStyleSheet("background-color: #3498db; color: white; font-weight: bold;")
        h_sel_1.addWidget(QLabel("매물ID:")); h_sel_1.addWidget(self.le_m_id); h_sel_1.addWidget(self.btn_load)
        h_sel_1.addStretch()
        v_sel_main.addLayout(h_sel_1)

        # [2행]: 주소/위치 통합 입력 (공통)
        h_sel_2 = QHBoxLayout()
        self.le_do = QLineEdit(); self.le_do.setPlaceholderText("도/특별시/광역시"); self.le_do.setFixedWidth(150)
        self.le_si_gun = QLineEdit(); self.le_si_gun.setPlaceholderText("시/군/구"); self.le_si_gun.setFixedWidth(140)
        self.le_emd_ri = QLineEdit(); self.le_emd_ri.setPlaceholderText("읍/면/동/리"); self.le_emd_ri.setFixedWidth(180)
        self.le_jibon = QLineEdit(); self.le_jibon.setPlaceholderText("지번"); self.le_jibon.setFixedWidth(100)
        self.chk_loc_public = QCheckBox("상세정보 공개"); self.chk_loc_public.setChecked(False)

        self.addr_fields = [self.le_do, self.le_si_gun, self.le_emd_ri, self.le_jibon]
        h_sel_2.addWidget(QLabel("주소(위치):"))
        for w in self.addr_fields: h_sel_2.addWidget(w)
        h_sel_2.addWidget(self.chk_loc_public)
        v_sel_main.addLayout(h_sel_2)

        # [3행]: 종류 선택 및 아파트 전용 세부 셀렉터
        h_sel_3 = QHBoxLayout()
        self.cb_type = QComboBox(); self.cb_type.addItems(self.category_config.keys())
        self.cb_type.currentTextChanged.connect(self.on_type_changed)
        
        self.lbl_reg = QLabel("지역:"); self.cb_region = QComboBox(); self.cb_region.setFixedWidth(130); self.cb_region.setEditable(True)
        self.lbl_cpx = QLabel("단지:"); self.cb_complex = QComboBox(); self.cb_complex.setFixedWidth(160); self.cb_complex.setEditable(True)
        self.lbl_dong = QLabel("동:"); self.cb_dong = QComboBox(); self.cb_dong.setFixedWidth(70); self.cb_dong.setEditable(True)
        self.lbl_ho = QLabel("호:"); self.cb_ho = QComboBox(); self.cb_ho.setFixedWidth(70); self.cb_ho.setEditable(True)

        # 글자색과 배경색을 명시적으로 지정하여 커서 시인성 확보
        apt_selection_widgets = [self.cb_region, self.cb_complex, self.cb_dong, self.cb_ho]
        for w in apt_selection_widgets:
            w.setStyleSheet("background-color: white; color: black; border: 1px solid #ced4da; border-radius: 4px;")

        self.cb_region.currentTextChanged.connect(self.update_complex_list)
        self.cb_complex.currentTextChanged.connect(self.update_dong_list)
        self.cb_dong.currentTextChanged.connect(self.update_ho_list)
        self.cb_ho.currentTextChanged.connect(self.auto_fill_apt_info)

        h_sel_3.addWidget(QLabel("매물종류:")); h_sel_3.addWidget(self.cb_type)
        for w in [self.lbl_reg, self.cb_region, self.lbl_cpx, self.cb_complex, self.lbl_dong, self.cb_dong, self.lbl_ho, self.cb_ho]:
            h_sel_3.addWidget(w)
        h_sel_3.addStretch()
        v_sel_main.addLayout(h_sel_3)

        # 초기화 데이터 로드
        self.cb_region.addItems(self.emd_ri_list)
        # 검색 편의성을 위해 포함(Contains) 검색 모드 적용
        if self.cb_region.completer():
            self.cb_region.completer().setFilterMode(Qt.MatchFlag.MatchContains)
        if self.cb_complex.completer():
            self.cb_complex.completer().setFilterMode(Qt.MatchFlag.MatchContains)
        self.update_complex_list()

        top_group.setLayout(v_sel_main)
        prop_layout.addWidget(top_group)

        self.entry_panel = QHBoxLayout()
        self.entry_panel.setSpacing(10)
        prop_layout.addLayout(self.entry_panel)
        
        self.factory_detail_group = self.create_factory_detail_ui()
        prop_layout.addWidget(self.factory_detail_group)

        self.btn_save = QPushButton("매물등록")
        self.chk_data_analysis = QCheckBox("데이터분석 및 저장")
        self.chk_data_analysis.setChecked(True)
        self.cb_sns_type = QComboBox(); self.cb_sns_type.addItems(["SNS문자 (개별)", "SNS문자 (다중)"])
        self.btn_sns = QPushButton("📱 SNS 문자 생성")

        self.btn_save.clicked.connect(self.process_save)
        self.btn_sns.clicked.connect(self.handle_sns_action)

        # 하단 버튼 및 위젯 멋있게 꾸미기
        self.btn_save.setFixedHeight(50)
        self.btn_save.setStyleSheet("background-color: #2c3e50; color: white; font-weight: bold; font-size: 16px; border-radius: 5px;")

        self.chk_data_analysis.setFixedHeight(50)
        self.chk_data_analysis.setStyleSheet("""
            QCheckBox {
                background-color: #34495e;
                color: white;
                font-weight: bold;
                font-size: 16px;
                border-radius: 5px;
                padding: 10px 15px;
                spacing: 10px;
            }
            QCheckBox::indicator {
                width: 20px;
                height: 20px;
                background-color: white;
                border-radius: 3px;
            }
            QCheckBox::indicator:checked {
                background-color: #2ecc71;
            }
        """)

        self.cb_sns_type.setFixedHeight(50)
        self.cb_sns_type.setStyleSheet("""
            QComboBox {
                background-color: #3498db;
                color: white;
                font-weight: bold;
                font-size: 16px;
                border-radius: 5px;
                padding-left: 10px;
            }
            QComboBox::drop-down { border: 0px; }
        """)

        self.btn_sns.setFixedHeight(50)
        self.btn_sns.setStyleSheet("background-color: #27ae60; color: white; font-weight: bold; font-size: 16px; border-radius: 5px;")

        # 하단 출력창 설정 (목록 확인을 위해 높이 확장)
        self.output = QTextEdit()
        self.output.setFixedHeight(180)
        self.output.setStyleSheet("border: 1px solid #ced4da; border-radius: 5px; background: white; font-size: 13px;")

        btn_h_layout = QHBoxLayout()
        btn_h_layout.addWidget(self.btn_save)
        btn_h_layout.addWidget(self.chk_data_analysis)
        btn_h_layout.addWidget(self.cb_sns_type)
        btn_h_layout.addWidget(self.btn_sns)

        prop_layout.addLayout(btn_h_layout)
        prop_layout.addWidget(self.output)

        # [스크롤 영역 추가] - 입력 내용이 길어질 경우 버튼이 사라지는 문제 해결
        prop_scroll = QScrollArea()
        prop_container = QWidget()
        prop_container.setLayout(prop_layout)
        prop_scroll.setWidget(prop_container)
        prop_scroll.setWidgetResizable(True)
        prop_scroll.setStyleSheet("border: none;")

        tab1_vbox = QVBoxLayout()
        tab1_vbox.addWidget(prop_scroll)
        self.property_tab.setLayout(tab1_vbox)

        # 2번 탭 초기화
        self.init_analysis_tab()
        self.init_marketing_tab()
        self.init_consult_tab()
        self.init_activity_tab()
        self.init_contract_tab()

        main_layout.addWidget(self.tabs)

        self.on_type_changed()
        self.setLayout(main_layout)

    def init_marketing_tab(self):
        """마케팅 대시보드 및 상담일지 UI 초기화"""
        layout = QVBoxLayout()
        layout.setSpacing(5) # 전체적인 레이아웃 간격 축소
        
        # [상단: 마케팅 대시보드 현황판]
        dash_group = QGroupBox("📌 매물 보유 현황")
        dash_group.setStyleSheet("QGroupBox { font-weight: bold; border: 2px solid #34495e; border-radius: 10px; margin-top: 10px; background: white; }")
        dash_layout = QHBoxLayout()
        dash_layout.setContentsMargins(10, 25, 10, 10) # 제목 공간 확보를 위해 상단 여백 25
        
        self.stat_labels = {}
        for cat in self.category_config.keys():
            v_box = QVBoxLayout(); v_box.setSpacing(0)
            title = QLabel(cat)
            title.setAlignment(Qt.AlignmentFlag.AlignCenter)
            title.setFixedHeight(30) # 행높이 30으로 설정
            val = QLabel("0건"); val.setStyleSheet("font-size: 12px; font-weight: bold; color: #2980b9;") # 폰트 크기 조정
            val.setAlignment(Qt.AlignmentFlag.AlignCenter)
            val.setFixedHeight(30) # 행높이 30으로 설정
            self.stat_labels[cat] = val
            v_box.addWidget(title)
            v_box.addWidget(val)
            dash_layout.addLayout(v_box)
        
        btn_refresh = QPushButton("🔄 통계 새로고침")
        btn_refresh.clicked.connect(self.refresh_dashboard)
        btn_refresh.setFixedWidth(120)
        btn_open_db = QPushButton("📊 엑셀 DB 열기")
        btn_open_db.clicked.connect(self.open_excel_db)
        btn_open_db.setFixedWidth(120)

        dash_layout.addWidget(btn_refresh)
        dash_layout.addWidget(btn_open_db)
        dash_group.setLayout(dash_layout)
        layout.addWidget(dash_group)

        # 하단 관련 버튼 (고객상담 탭에서 이전됨)
        btn_generate_urgent = QPushButton("🚀 관심고객용 [급매물] SNS 문구 생성")
        btn_generate_urgent.setStyleSheet("background-color: #e74c3c; color: white; font-weight: bold; height: 35px;")
        btn_generate_urgent.clicked.connect(self.generate_urgent_sns_text)

        self.btn_export_blog = QPushButton("🌐 블로그용 공개 매물장 추출 (엑셀)")
        self.btn_export_blog.setStyleSheet("background-color: #27ae60; color: white; font-weight: bold; height: 35px;")
        self.btn_export_blog.clicked.connect(self.export_for_blog)

        btn_box = QHBoxLayout()
        btn_box.addWidget(btn_generate_urgent)
        btn_box.addWidget(self.btn_export_blog)
        layout.addLayout(btn_box)

        layout.addStretch() # 대시보드를 상단으로 밀어올림
        
        # 마케팅 탭에도 스크롤 적용
        mark_scroll = QScrollArea()
        mark_container = QWidget()
        mark_container.setLayout(layout)
        mark_scroll.setWidget(mark_container)
        mark_scroll.setWidgetResizable(True)
        mark_scroll.setStyleSheet("border: none;")

        tab2_vbox = QVBoxLayout()
        tab2_vbox.addWidget(mark_scroll)
        self.marketing_tab.setLayout(tab2_vbox)

    def init_analysis_tab(self):
        """데이터 분석 탭 UI 초기화"""
        layout = QVBoxLayout()
        label = QLabel("📈 데이터 분석 기능 준비 중입니다.")
        label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(label)
        self.analysis_tab.setLayout(layout)

    def init_consult_tab(self):
        """고객상담 UI 초기화 - 마케팅 탭에서 이전됨"""
        layout = QVBoxLayout()
        layout.setSpacing(10)
        layout.setContentsMargins(10, 10, 10, 10)

        # [상담일지 기록 섹션]
        consult_group = QGroupBox("📝 고객 상담일지 작성")
        consult_group.setStyleSheet("QGroupBox { font-weight: bold; border: 1px solid #ced4da; background: #fdfefe; margin-top: 10px; }")
        
        # 하단 여백 삭제를 위해 그룹박스 내부 레이아웃 구성
        consult_vbox = QVBoxLayout()
        consult_vbox.setContentsMargins(15, 25, 15, 15)
        consult_vbox.setSpacing(10)

        c_layout = QGridLayout(); c_layout.setVerticalSpacing(10)

        self.le_client_name = QLineEdit(); self.le_client_name.setPlaceholderText("고객명")
        self.le_client_contact = QLineEdit(); self.le_client_contact.setPlaceholderText("연락처")
        self.cb_target_type = QComboBox(); self.cb_target_type.addItems(list(self.category_config.keys())); self.cb_target_type.setFixedWidth(120)
        self.le_target_id = QLineEdit(); self.le_target_id.setPlaceholderText("매물 ID (예: 260528-01)")
        self.cb_consult_type = QComboBox(); self.cb_consult_type.addItems(["WI", "전화"])
        self.te_consult_memo = QTextEdit(); self.te_consult_memo.setPlaceholderText("상담 상세 내용을 입력하세요. (요청사항, 제한사항, 예산 등)")
        self.te_consult_memo.setFixedHeight(240)
        
        c_layout.addWidget(QLabel("고객정보:"), 0, 0); c_layout.addWidget(self.le_client_name, 0, 1, 1, 2); c_layout.addWidget(self.le_client_contact, 0, 3, 1, 2)
        c_layout.addWidget(QLabel("관심매물:"), 1, 0); c_layout.addWidget(self.cb_target_type, 1, 1); c_layout.addWidget(self.le_target_id, 1, 2)
        c_layout.addWidget(QLabel("방문:"), 1, 3); c_layout.addWidget(self.cb_consult_type, 1, 4)
        c_layout.addWidget(QLabel("상담내용:"), 2, 0); c_layout.addWidget(self.te_consult_memo, 2, 1, 1, 4)
        
        consult_vbox.addLayout(c_layout)
        consult_group.setLayout(consult_vbox)
        layout.addWidget(consult_group)

        # 하단 버튼 박스
        # 하단 관련 버튼
        btn_save_consult = QPushButton("💾 상담내용 DB 저장")
        btn_save_consult.setStyleSheet("background-color: #8e44ad; color: white; font-weight: bold; height: 35px;")
        btn_save_consult.clicked.connect(self.save_consultation_data)

        btn_box = QHBoxLayout()
        btn_box.addWidget(btn_save_consult)
        
        layout.addLayout(btn_box)
        layout.addStretch()

        # 상담 탭에도 스크롤 적용
        cons_scroll = QScrollArea()
        cons_container = QWidget()
        cons_container.setLayout(layout)
        cons_scroll.setWidget(cons_container)
        cons_scroll.setWidgetResizable(True)
        cons_scroll.setStyleSheet("border: none;")

        tab_vbox = QVBoxLayout()
        tab_vbox.addWidget(cons_scroll)
        self.consult_tab.setLayout(tab_vbox)

    def init_activity_tab(self):
        """임장활동 UI 초기화"""
        layout = QVBoxLayout()
        layout.setSpacing(15)
        layout.setContentsMargins(10, 10, 10, 10)

        # [1단계: 준비단계]
        prep_group = QGroupBox("📋 1. 준비단계 (매물 선정 및 확인)")
        prep_group.setStyleSheet("QGroupBox { font-weight: bold; border: 1px solid #ced4da; background: #fdfefe; margin-top: 10px; }")
        prep_layout = QGridLayout()
        prep_layout.setContentsMargins(15, 25, 15, 15)
        prep_layout.setVerticalSpacing(10)
        self.te_recommend_list = QTextEdit(); self.te_recommend_list.setPlaceholderText("중개사 추천매물 List up (매물 ID 또는 간략 정보)"); self.te_recommend_list.setFixedHeight(80)
        self.le_view_pw = QLineEdit(); self.le_view_pw.setPlaceholderText("비밀번호 (공실)")
        self.chk_view_confirmed = QCheckBox("집보기 예약 확인")
        self.le_owner_contact_act = QLineEdit(); self.le_owner_contact_act.setPlaceholderText("소유자 연락처 확인")
        
        prep_layout.addWidget(QLabel("추천매물 리스트:"), 0, 0)
        prep_layout.addWidget(self.te_recommend_list, 0, 1, 1, 2)
        prep_layout.addWidget(QLabel("공실 비번:"), 1, 0)
        prep_layout.addWidget(self.le_view_pw, 1, 1)
        prep_layout.addWidget(self.chk_view_confirmed, 1, 2)
        prep_layout.addWidget(QLabel("소유자 연락처:"), 2, 0)
        prep_layout.addWidget(self.le_owner_contact_act, 2, 1, 1, 2)
        prep_group.setLayout(prep_layout)
        layout.addWidget(prep_group)

        # [2단계: 임장활동]
        activity_group = QGroupBox("🚶 2. 임장활동 (현장 방문 계획)")
        activity_group.setStyleSheet("QGroupBox { font-weight: bold; border: 1px solid #ced4da; background: #f4f6f7; margin-top: 10px; }")
        activity_grid = QGridLayout()
        activity_grid.setContentsMargins(15, 25, 15, 15)
        activity_grid.setVerticalSpacing(10)
        
        self.le_route_plan = QLineEdit(); self.le_route_plan.setPlaceholderText("동선 계획")
        self.le_time_plan = QLineEdit(); self.le_time_plan.setPlaceholderText("시간 계획")
        
        activity_grid.addWidget(QLabel("동선 계획:"), 0, 0)
        activity_grid.addWidget(self.le_route_plan, 0, 1)
        activity_grid.addWidget(QLabel("시간 계획:"), 1, 0)
        activity_grid.addWidget(self.le_time_plan, 1, 1)
        activity_group.setLayout(activity_grid)
        layout.addWidget(activity_group)

        # [3단계: 임장후기 정리]
        review_group = QGroupBox("📝 3. 임장후기 정리 (피드백 관리)")
        review_group.setStyleSheet("QGroupBox { font-weight: bold; border: 1px solid #ced4da; background: #fffde7; margin-top: 10px; }")
        review_layout = QGridLayout()
        review_layout.setContentsMargins(15, 25, 15, 15)
        review_layout.setVerticalSpacing(10)
        
        self.te_activity_list_mgmt = QTextEdit(); self.te_activity_list_mgmt.setPlaceholderText("임장 List 관리"); self.te_activity_list_mgmt.setFixedHeight(80)
        self.le_customer_reaction = QLineEdit(); self.le_customer_reaction.setPlaceholderText("고객 반응 (선호도 등)")
        self.le_hope_price = QLineEdit(); self.le_hope_price.setPlaceholderText("고객 희망 가격")
        self.te_prop_changes = QTextEdit(); self.te_prop_changes.setPlaceholderText("임장 시 매물 변경사항 체크정리 (수선 필요, 옵션 변동 등)"); self.te_prop_changes.setFixedHeight(80)
        
        review_layout.addWidget(QLabel("임장 List 관리:"), 0, 0)
        review_layout.addWidget(self.te_activity_list_mgmt, 0, 1, 1, 3)
        review_layout.addWidget(QLabel("고객 반응:"), 1, 0)
        review_layout.addWidget(self.le_customer_reaction, 1, 1)
        review_layout.addWidget(QLabel("희망 가격:"), 1, 2)
        review_layout.addWidget(self.le_hope_price, 1, 3)
        review_layout.addWidget(QLabel("변경사항 정리:"), 2, 0)
        review_layout.addWidget(self.te_prop_changes, 2, 1, 1, 3)
        review_group.setLayout(review_layout)
        layout.addWidget(review_group)

        layout.addStretch() # 아래쪽 여백 확보

        # 스크롤 영역 적용
        act_scroll = QScrollArea()
        act_container = QWidget()
        act_container.setLayout(layout)
        act_scroll.setWidget(act_container)
        act_scroll.setWidgetResizable(True)
        act_scroll.setStyleSheet("border: none;")

        tab_vbox = QVBoxLayout()
        tab_vbox.addWidget(act_scroll)
        self.activity_tab.setLayout(tab_vbox)

    def init_contract_tab(self):
        """계약 관리 탭 UI 초기화"""
        layout = QVBoxLayout()
        label = QLabel("📄 계약 관리 기능 준비 중입니다.")
        label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(label)
        self.contract_tab.setLayout(layout)

    def load_apt_master_db(self):
        """아파트 정보를 Excel 파일에서 로드하여 드롭다운용 데이터 구성 (apt_complex_info.xlsx)"""
        # Excel 파일에서 로드하여 드롭다운용 데이터 구성
        self.apt_master_db = {}
        self.emd_ri_to_full_reg_map = {}
        self.complex_dong_ho_map = {}
        self.apt_unit_details = {}
        self.emd_ri_list = []

        excel_path = Config.APT_MASTER_EXCEL_PATH
        if not os.path.exists(excel_path):
            self.output.append(f"⚠️ 경고: {excel_path} 파일을 찾을 수 없습니다. 아파트 마스터 데이터가 없어 자동 완성 기능이 제한됩니다.")
            return

        try:
            # Load all sheets from the Excel file
            excel_data = pd.read_excel(excel_path, sheet_name=None, engine='openpyxl')

            # 1. Process 'apt_complex_info' sheet
            if 'apt_complex_info' in excel_data:
                df_info = excel_data['apt_complex_info'].fillna('')
                for _, item in df_info.iterrows():
                    reg = str(item.get('지역', '미분류'))
                    emd_ri = str(item.get('읍면동리', ''))
                    c_name = str(item.get('단지명', ''))
                    if emd_ri:
                        if emd_ri not in self.emd_ri_to_full_reg_map: self.emd_ri_to_full_reg_map[emd_ri] = set()
                        self.emd_ri_to_full_reg_map[emd_ri].add(reg)
                    if reg not in self.apt_master_db: self.apt_master_db[reg] = set()
                    self.apt_master_db[reg].add(c_name)
                self.emd_ri_list = sorted(list(self.emd_ri_to_full_reg_map.keys()))

            # 2. Process 'apt_complex_floor' sheet
            if 'apt_complex_floor' in excel_data:
                df_floor = excel_data['apt_complex_floor'].fillna('')
                for _, item in df_floor.iterrows():
                    c_name = str(item.get('단지명', ''))
                    dong = str(item.get('동', ''))
                    ho = str(item.get('호', ''))
                    if c_name not in self.complex_dong_ho_map: self.complex_dong_ho_map[c_name] = {}
                    if dong not in self.complex_dong_ho_map[c_name]: self.complex_dong_ho_map[c_name][dong] = []
                    if ho: self.complex_dong_ho_map[c_name][dong].append(ho)

            # 3. Process 'apt_complex_cost' sheet
            if 'apt_complex_cost' in excel_data:
                df_cost = excel_data['apt_complex_cost'].fillna('')
                for _, item in df_cost.iterrows():
                    c_name = str(item.get('단지명', ''))
                    dong = str(item.get('동', ''))
                    ho = str(item.get('호', ''))
                    self.apt_unit_details[(c_name, dong, ho)] = {
                        "type": str(item.get('타입', '-')),
                        "supp": str(item.get('공급면적', '-')),
                        "priv": str(item.get('전용면적', '-'))
                    }
        except Exception as e:
            print(f"아파트 데이터 로드 중 오류: {e}")
            QMessageBox.critical(self, "데이터 로드 오류", f"아파트 마스터 데이터 로드 중 오류 발생: {e}\n'apt_complex_info.xlsx' 파일이 올바른 형식인지 확인해주세요.")

    def update_complex_list(self):
        selected_emd_ri = self.cb_region.currentText()
        self.cb_complex.clear()
        if selected_emd_ri in self.emd_ri_to_full_reg_map:
            complexes = set()
            for full_reg in self.emd_ri_to_full_reg_map[selected_emd_ri]:
                if full_reg in self.apt_master_db:
                    complexes.update(self.apt_master_db[full_reg])
            self.cb_complex.addItems(sorted(list(complexes)))
            # 지역이 정확히 입력되면 단지 목록 드롭다운을 자동으로 실행합니다.
            if complexes:
                self.cb_complex.showPopup()
        self.update_dong_list()

    def update_dong_list(self):
        c_name = self.cb_complex.currentText()
        self.cb_dong.clear()
        if c_name in self.complex_dong_ho_map:
            dongs = self.complex_dong_ho_map[c_name].keys()
            self.cb_dong.addItems(sorted(list(dongs)))
        self.update_ho_list()

    def update_ho_list(self):
        c_name = self.cb_complex.currentText()
        dong = self.cb_dong.currentText()
        self.cb_ho.clear()
        if c_name in self.complex_dong_ho_map and dong in self.complex_dong_ho_map[c_name]:
            hos = self.complex_dong_ho_map[c_name][dong]
            self.cb_ho.addItems(sorted(list(set(hos))))
        self.auto_fill_apt_info()

    def auto_fill_apt_info(self):
        """주소 자동 보정 및 G2 구역 면적 데이터 자동 입력 + DB 자동 업데이트"""
        selected_emd_ri = self.cb_region.currentText()
        c_name = self.cb_complex.currentText()
        dong = self.cb_dong.currentText()
        ho = self.cb_ho.currentText().strip()

        if selected_emd_ri:
            self.le_emd_ri.setText(selected_emd_ri)

        info = self.apt_unit_details.get((c_name, dong, ho))

        if info:
            g2 = self.all_fields.get("G2", {})
            if "공급면적(㎡)" in g2: g2["공급면적(㎡)"].setText(info.get('supp', '-'))
            if "전용면적(㎡)" in g2: g2["전용면적(㎡)"].setText(info.get('priv', '-'))

            # 지역(동), 단지명, 동, 호가 모두 입력되었을 때 estate_db.xlsx 자동 업데이트
            if all([selected_emd_ri, c_name, dong, ho]) and self.cb_type.currentText() == "아파트":
                update_data = {
                    "공급면적": info.get('supp', '-'),
                    "전용면적": info.get('priv', '-'),
                    "타입": info.get('type', '-')
                }
                update_apartment_in_excel(self.db_path, c_name, dong, ho, update_data)

    def refresh_dashboard(self):
        """엑셀 DB를 읽어 현재 매물 보유 현황을 대시보드에 반영"""
        if not os.path.exists(self.db_path): return
        try:
            with pd.ExcelFile(self.db_path) as reader:
                for cat in self.category_config.keys():
                    if cat in reader.sheet_names:
                        df = pd.read_excel(reader, sheet_name=cat)
                        count = len(df)
                        self.stat_labels[cat].setText(f"{count}건")
        except Exception as e:
            self.output.append(f"📊 통계 새로고침 오류: {e}")

    def open_excel_db(self):
        """엑셀 DB 파일을 열기"""
        if not os.path.exists(self.db_path):
            QMessageBox.warning(self, "오류", "데이터베이스 파일이 존재하지 않습니다.")
            return
        try:
            os.startfile(self.db_path) # Windows에서 파일 열기
        except Exception as e:
            QMessageBox.critical(self, "오류", f"엑셀 파일을 열 수 없습니다.\n원인: {str(e)}")

    def save_consultation_data(self):
        """상담일지를 별도의 '상담일지' 시트에 저장"""
        client = self.le_client_name.text().strip()
        if not client: 
            QMessageBox.warning(self, "알림", "고객명을 입력해주세요.")
            return
            
        record = {
            "상담일자": datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
            "고객명": client,
            "연락처": self.le_client_contact.text(),
            "관심매물ID": f"{self.cb_target_type.currentText()}-{self.le_target_id.text()}", # 매물 종류와 ID 조합
            "상담구분": self.cb_consult_type.currentText(),
            "상담내용": self.te_consult_memo.toPlainText()
        }
        
        # 엑셀 저장 로직 (상담일지 시트)
        try:
            df_new = pd.DataFrame([record])
            if os.path.exists(self.db_path):
                with pd.ExcelWriter(self.db_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    if '상담일지' in writer.book.sheetnames:
                        start_row = writer.book['상담일지'].max_row
                        df_new.to_excel(writer, sheet_name='상담일지', index=False, header=False, startrow=start_row)
                    else:
                        df_new.to_excel(writer, sheet_name='상담일지', index=False)
            else:
                df_new.to_excel(self.db_path, sheet_name='상담일지', index=False)
            
            QMessageBox.information(self, "저장 완료", f"{client} 고객님의 상담일지가 저장되었습니다.")
            self.le_client_name.clear(); self.le_client_contact.clear(); self.te_consult_memo.clear()
        except Exception as e:
            QMessageBox.critical(self, "오류", f"상담일지 저장 실패: {str(e)}")

    def generate_urgent_sns_text(self):
        """관심 고객을 위한 급매물 특화 SNS 문구 생성"""
        target_id_text = self.le_target_id.text().strip()
        if not target_id_text:
            QMessageBox.warning(self, "알림", "관심매물 ID를 입력한 후 문구를 생성해주세요.")
            return
        m_id = f"{self.cb_target_type.currentText()}-{target_id_text}" # 매물 종류와 ID 조합

        text = f"""[성균관 부동산 추천매물]
━━━━━━━━━━━━━━
⭐ 특별 관리 매물 안내 ⭐

고객님께서 관심 가지실 만한 
특별한 [급매/우수] 매물이 접수되었습니다!

🆔 매물번호: {m_id}
📍 상세내용: 지금 바로 확인하세요.

💰 특징: 가격 조정 완료 및 로열층 확보!

상세 내역은 문의주시면 
친절하게 안내해 드리겠습니다.\n{Config.SNS_FOOTER}"""
        self.output.clear()
        self.output.append(text)
        self.tabs.setCurrentIndex(0) # 텍스트 확인을 위해 매물 탭으로 이동
        QMessageBox.information(self, "생성 완료", "추천 문구가 생성되었습니다. 하단 출력창을 확인하세요.")

    def export_for_blog(self):
        """비공개 정보를 제외하고 공개 정보만 필터링하여 블로그용 엑셀 생성"""
        if not os.path.exists(self.db_path):
            QMessageBox.warning(self, "오류", "데이터베이스 파일이 없습니다.")
            return

        export_path = "매물장_블로그공개용.xlsx"
        try:
            with pd.ExcelFile(self.db_path, engine='openpyxl') as reader:
                with pd.ExcelWriter(export_path, engine='openpyxl') as writer:
                    for sheet_name in reader.sheet_names:
                        if sheet_name == "상담일지": continue # 상담일지는 제외
                        
                        df = pd.read_excel(reader, sheet_name=sheet_name)
                        if df.empty: continue

                        # 비공개 구역(지번, 호, 소유자 등) 시작 지점 찾기
                        # 우리 시스템의 규칙: "미디어경로"까지가 공개 정보임
                        if "미디어경로(사진/영상)" in df.columns:
                            media_idx = list(df.columns).index("미디어경로(사진/영상)")
                            # 미디어경로까지만 포함하고 그 뒤(지번, 소유자 등)는 삭제
                            df_public = df.iloc[:, :media_idx + 1]
                        else:
                            # 미디어경로 컬럼이 없는 경우 안전을 위해 지번/소유자 수동 체크
                            private_triggers = ["지번", "호", "소유자", "연락처"]
                            idx = len(df.columns)
                            for trigger in private_triggers:
                                if trigger in df.columns:
                                    idx = min(idx, list(df.columns).index(trigger))
                            df_public = df.iloc[:, :idx]

                        df_public.to_excel(writer, sheet_name=sheet_name, index=False)
            
            QMessageBox.information(self, "추출 완료", f"공개용 파일이 생성되었습니다.\n경로: {os.path.abspath(export_path)}\n\n이 파일을 구글 시트에 업로드하여 블로그에 게시하세요.")
        except Exception as e:
            QMessageBox.critical(self, "오류", f"추출 중 에러 발생: {str(e)}")

    def select_media(self):
        """파일 다이얼로그를 통해 사진 및 영상 선택"""
        files, _ = QFileDialog.getOpenFileNames(self, "매물 사진/영상 선택", "", "Media Files (*.png *.jpg *.jpeg *.gif *.mp4 *.mov *.avi *.mkv)")
        if files:
            self.selected_media = files
            if hasattr(self, 'btn_add_media'):
                self.btn_add_media.setText(f"사진,영상링크 ({len(self.selected_media)})")

    def create_commercial_detail_ui(self):
        """상가 상세 내역 입력을 위한 UI 그룹 생성"""
        group = QGroupBox("상가접수 상세 (층별/호실별 현황/단위:㎡/만원)")
        # margin-top을 30px로 늘려 상단 박스들과의 겹침 방지
        group.setStyleSheet("QGroupBox { font-weight: bold; border: 1px solid #ced4da; margin-top: 30px; }")
        layout = QVBoxLayout()
        layout.setContentsMargins(15, 35, 15, 15) # 제목 영역 확보를 위해 상단 내부 여백 확대
        layout.setSpacing(10)

        # 테이블을 먼저 배치
        self.table = QTableWidget(0, 10)
        self.table.setHorizontalHeaderLabels(["층", "호실", "용도", "면적", "전세보증금", "월세보증금", "월세", "옵션", "현재상태", "임차만기일"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setFixedHeight(200)
        self.table.itemChanged.connect(self.on_table_item_changed)

        # 행 추가/삭제 버튼 레이아웃 (테이블 아래로 이동)
        btn_layout = QHBoxLayout()
        self.btn_add_row = QPushButton("+ 행 추가")
        self.btn_del_row = QPushButton("- 행 삭제")
        self.btn_add_row.clicked.connect(self.add_detail_row)
        self.btn_del_row.clicked.connect(self.del_detail_row)
        btn_layout.addWidget(self.btn_add_row)
        btn_layout.addWidget(self.btn_del_row)
        btn_layout.addStretch()
        
        self.lbl_totals = QLabel("합계: 보증금 0만원 / 월세 0만원 (연 0만원)")
        self.lbl_totals.setStyleSheet("font-weight: bold; color: blue; font-size: 14px;")
        
        layout.addWidget(self.table)
        layout.addLayout(btn_layout)
        layout.addWidget(self.lbl_totals)
        group.setLayout(layout)
        group.setVisible(False)
        return group

    def create_house_detail_ui(self):
        """주택 상세 내역 입력을 위한 UI 그룹 생성 (상가와 동일 양식)"""
        group = QGroupBox("주택접수 상세 (층별/호실별 현황/단위:㎡/만원)")
        group.setStyleSheet("QGroupBox { font-weight: bold; border: 1px solid #ced4da; margin-top: 30px; }")
        layout = QVBoxLayout()
        layout.setContentsMargins(15, 35, 15, 15)
        layout.setSpacing(10)

        self.table_house = QTableWidget(0, 10)
        self.table_house.setHorizontalHeaderLabels(["층", "호실", "용도", "면적", "전세보증금", "월세보증금", "월세", "옵션", "현재상태", "임차만기일"])
        self.table_house.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.table_house.horizontalHeader().setStretchLastSection(True)
        self.table_house.setFixedHeight(200)
        self.table_house.itemChanged.connect(self.on_house_table_item_changed)

        btn_layout = QHBoxLayout()
        self.btn_add_row_h = QPushButton("+ 행 추가")
        self.btn_del_row_h = QPushButton("- 행 삭제")
        self.btn_add_row_h.clicked.connect(self.add_house_row)
        self.btn_del_row_h.clicked.connect(self.del_house_row)
        btn_layout.addWidget(self.btn_add_row_h)
        btn_layout.addWidget(self.btn_del_row_h)
        btn_layout.addStretch()
        
        self.lbl_house_totals = QLabel("합계: 보증금 0만원 / 월세 0만원 (연 0만원)")
        self.lbl_house_totals.setStyleSheet("font-weight: bold; color: #e67e22; font-size: 14px;")
        
        layout.addWidget(self.table_house)
        layout.addLayout(btn_layout)
        layout.addWidget(self.lbl_house_totals)
        group.setLayout(layout)
        group.setVisible(False)
        return group

    def add_house_row(self): self.table_house.insertRow(self.table_house.rowCount())
    def del_house_row(self):
        curr = self.table_house.currentRow()
        if curr >= 0: self.table_house.removeRow(curr)
    def create_land_detail_ui(self):
        """토지 상세 내역(여러 필지) 입력을 위한 UI 그룹 생성"""
        group = QGroupBox("토지접수 상세 (필지별 현황/단위:㎡/만원)")
        group.setStyleSheet("QGroupBox { font-weight: bold; border: 1px solid #ced4da; margin-top: 30px; }")
        layout = QVBoxLayout()
        layout.setContentsMargins(15, 35, 15, 15)
        layout.setSpacing(10)

        self.table_land = QTableWidget(0, 11)
        self.table_land.setHorizontalHeaderLabels(["필지ID", "도/특별시/광역시", "시/군/구", "읍/면/동/리", "지번", "용도지역", "지목", "면적(㎡)", "개별공시지가(㎡,원)", "매매가(만원)", "소유자"])
        self.table_land.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.table_land.horizontalHeader().setStretchLastSection(True)
        self.table_land.setFixedHeight(200)
        self.table_land.itemChanged.connect(self.on_land_table_item_changed)

        btn_layout = QHBoxLayout()
        self.btn_add_row_l = QPushButton("+ 필지 추가")
        self.btn_del_row_l = QPushButton("- 필지 삭제")
        self.btn_add_row_l.clicked.connect(self.add_land_row)
        self.btn_del_row_l.clicked.connect(self.del_land_row)
        btn_layout.addWidget(self.btn_add_row_l)
        btn_layout.addWidget(self.btn_del_row_l)
        btn_layout.addStretch()
        
        self.lbl_land_totals = QLabel("합계: 총 면적 0㎡ / 총 매매가 0만원")
        self.lbl_land_totals.setStyleSheet("font-weight: bold; color: #27ae60; font-size: 14px;")
        
        layout.addWidget(self.table_land)
        layout.addLayout(btn_layout)
        layout.addWidget(self.lbl_land_totals)
        group.setLayout(layout)
        group.setVisible(False)
        return group

    def create_factory_detail_ui(self):
        """공장/창고 상세 내역 입력을 위한 UI 그룹 생성"""
        group = QGroupBox("공장/창고 상세 (층별 현황/단위:㎡/만원)")
        group.setStyleSheet("QGroupBox { font-weight: bold; border: 1px solid #ced4da; margin-top: 30px; }")
        layout = QVBoxLayout()
        layout.setContentsMargins(15, 35, 15, 15)
        layout.setSpacing(10)

        self.table_factory = QTableWidget(0, 7)
        self.table_factory.setHorizontalHeaderLabels(["층", "구조", "용도", "면적(㎡)", "보증금(만원)", "월세(만원)", "비고"])
        self.table_factory.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.table_factory.horizontalHeader().setStretchLastSection(True)
        self.table_factory.setFixedHeight(200)
        self.table_factory.itemChanged.connect(self.on_factory_table_item_changed)

        btn_layout = QHBoxLayout()
        self.btn_add_row_f = QPushButton("+ 층 추가")
        self.btn_del_row_f = QPushButton("- 층 삭제")
        self.btn_add_row_f.clicked.connect(self.add_factory_row)
        self.btn_del_row_f.clicked.connect(self.del_factory_row)
        btn_layout.addWidget(self.btn_add_row_f)
        btn_layout.addWidget(self.btn_del_row_f)
        btn_layout.addStretch()
        
        self.lbl_factory_totals = QLabel("합계: 보증금 0만원 / 월세 0만원 (수익률: 0.00%)")
        self.lbl_factory_totals.setStyleSheet("font-weight: bold; color: #8e44ad; font-size: 14px;")
        
        layout.addWidget(self.table_factory)
        layout.addLayout(btn_layout)
        layout.addWidget(self.lbl_factory_totals)
        group.setLayout(layout)
        group.setVisible(False)
        return group

    def add_factory_row(self):
        self.table_factory.insertRow(self.table_factory.rowCount())

    def del_factory_row(self):
        curr = self.table_factory.currentRow()
        if curr >= 0: self.table_factory.removeRow(curr)

    def add_land_row(self):
        self.table_land.insertRow(self.table_land.rowCount())

    def del_land_row(self):
        curr = self.table_land.currentRow()
        if curr >= 0: self.table_land.removeRow(curr)

    def add_detail_row(self):
        self.table.insertRow(self.table.rowCount())

    def del_detail_row(self):
        curr = self.table.currentRow()
        if curr >= 0: self.table.removeRow(curr)

    def on_table_item_changed(self, item):
        """테이블 셀 수정 시 숫자 포맷팅 및 합계 계산 호출"""
        col = item.column()
        # 면적(3), 전세보증금(4), 월세보증금(5), 월세(6) 컬럼에 대해 콤마 처리
        if col in [3, 4, 5, 6]:
            text = item.text().replace(",", "").strip()
            if text.isdigit():
                formatted = format(int(text), ",")
                if item.text() != formatted:
                    self.table.blockSignals(True)
                    item.setText(formatted)
                    self.table.blockSignals(False)
        self.calculate_commercial_totals()

    def on_house_table_item_changed(self, item):
        col = item.column()
        if col in [3, 4, 5, 6]:
            text = item.text().replace(",", "").strip()
            if text.isdigit():
                formatted = format(int(text), ",")
                if item.text() != formatted:
                    self.table_house.blockSignals(True)
                    item.setText(formatted)
                    self.table_house.blockSignals(False)
        self.calculate_house_totals()

    def on_land_table_item_changed(self, item):
        """토지 테이블 수정 시 숫자 포맷팅 및 합계 계산"""
        col = item.column()
        # 7: 면적, 8: 개별공시지가, 9: 매매가
        if col in [7, 8, 9]:
            text = item.text().replace(",", "").replace(" ", "").strip()
            if text.isdigit():
                formatted = format(int(text), ",")
                if item.text() != formatted:
                    self.table_land.blockSignals(True)
                    item.setText(formatted)
                    self.table_land.blockSignals(False)
        self.calculate_land_totals()

    def on_factory_table_item_changed(self, item):
        """공장 테이블 수정 시 숫자 포맷팅 및 합계 계산"""
        col = item.column()
        # 3: 면적, 4: 보증금, 5: 월세
        if col in [3, 4, 5]:
            text = item.text().replace(",", "").strip()
            if text.isdigit():
                formatted = format(int(text), ",")
                if item.text() != formatted:
                    self.table_factory.blockSignals(True)
                    item.setText(formatted)
                    self.table_factory.blockSignals(False)
        self.calculate_factory_totals()

    def calculate_factory_totals(self):
        """공장 테이블의 보증금/월세 합계 및 수익률 계산"""
        total_dep = 0
        total_mon = 0
        
        price_field = self.all_fields.get("G3", {}).get("매매가")
        price_text = price_field.text().replace(",", "") if price_field else ""
        price_val = int(price_text) if price_text.isdigit() else 0

        for r in range(self.table_factory.rowCount() if hasattr(self, 'table_factory') else 0):
            try:
                dep_item = self.table_factory.item(r, 4)
                mon_item = self.table_factory.item(r, 5)
                if dep_item and dep_item.text().replace(',', '').strip().isdigit():
                    total_dep += int(dep_item.text().replace(',', '').strip())
                if mon_item and mon_item.text().replace(',', '').strip().isdigit():
                    total_mon += int(mon_item.text().replace(',', '').strip())
            except (ValueError, TypeError):
                continue
        
        total_yearly = total_mon * 12
        yield_rate = 0
        investment = price_val - total_dep
        if investment > 0:
            yield_rate = (total_yearly / investment) * 100

        self.lbl_factory_totals.setText(f"합계: 보증금 {total_dep:,}만원 / 월세 {total_mon:,}만원 (연 {total_yearly:,}만원) 수익률: {yield_rate:.2f}%")

    def calculate_land_totals(self):
        """토지 테이블의 총 면적과 총 매매가 합계를 계산"""
        total_area = 0
        total_price = 0
        for r in range(self.table_land.rowCount()):
            try:
                area_item = self.table_land.item(r, 7)
                price_item = self.table_land.item(r, 9)
                if area_item and area_item.text().replace(',', '').strip().isdigit():
                    total_area += int(area_item.text().replace(',', '').strip())
                if price_item and price_item.text().replace(',', '').strip().isdigit():
                    total_price += int(price_item.text().replace(',', '').strip())
            except (ValueError, TypeError):
                continue
            
        # 상단 메인 필드와 자동 동기화
        area_field = self.all_fields.get("G2", {}).get("면적(㎡)")
        price_field = self.all_fields.get("G3", {}).get("매매가")
        
        if area_field and total_area > 0:
            area_field.blockSignals(True)
            area_field.setText(format(total_area, ","))
            area_field.blockSignals(False)
        if price_field and total_price > 0:
            price_field.blockSignals(True)
            price_field.setText(format(total_price, ","))
            price_field.blockSignals(False)
            
        self.lbl_land_totals.setText(f"합계: 총 면적 {total_area:,}㎡ / 총 매매가 {total_price:,}만원")

    def calculate_commercial_totals(self):
        """테이블의 보증금과 월세 합계를 계산"""
        total_deposit = 0
        total_monthly = 0
        
        # 매매가 가져오기 (단위: 만원)
        price_field = self.all_fields.get("G3", {}).get("희망매매가(만원)")
        price_text = price_field.text().replace(",", "") if price_field else ""
        price_val = int(price_text) if price_text.isdigit() else 0

        for r in range(self.table.rowCount()):
            try:
                # 4: 전세보증금, 5: 월세보증금, 6: 월세
                jeonse_item = self.table.item(r, 4)
                w_dep_item = self.table.item(r, 5)
                mon_item = self.table.item(r, 6)
                
                # 보증금 합산 (전세 + 월세보증금)
                for item in [jeonse_item, w_dep_item]:
                    if item and item.text().replace(',', '').strip().isdigit():
                        total_deposit += int(item.text().replace(',', '').strip())
                
                # 월세 합산
                if mon_item and mon_item.text().replace(',', '').isdigit():
                    total_monthly += int(mon_item.text().replace(',', ''))
            except (ValueError, TypeError):
                continue
        
        total_yearly = total_monthly * 12
        
        # 수익률 계산 (취득세 등 비용 제외 순수 수익률)
        yield_rate = 0
        investment = price_val - total_deposit
        if investment > 0:
            yield_rate = (total_yearly / investment) * 100

        # 수익율(%) 입력 필드에도 자동 반영
        yield_field = self.all_fields.get("G3", {}).get("수익율(%)")
        if yield_field:
            yield_field.setText(f"{yield_rate:.2f}")

        self.lbl_totals.setText(f"합계: 보증금 {total_deposit:,}만원 / 월세 {total_monthly:,}만원 (연 {total_yearly:,}만원) 수익률: {yield_rate:.2f}%")

    def calculate_house_totals(self):
        total_deposit = 0
        total_monthly = 0
        for r in range(self.table_house.rowCount()):
            try:
                jeonse_item = self.table_house.item(r, 4)
                w_dep_item = self.table_house.item(r, 5)
                mon_item = self.table_house.item(r, 6)
                for item in [jeonse_item, w_dep_item]:
                    if item and item.text().replace(',', '').strip().isdigit():
                        total_deposit += int(item.text().replace(',', '').strip())
                if mon_item and mon_item.text().replace(',', '').isdigit():
                    total_monthly += int(mon_item.text().replace(',', ''))
            except (ValueError, TypeError):
                continue
        
        total_yearly = total_monthly * 12
        price_field = self.all_fields.get("G3", {}).get("매매가")
        price_text = price_field.text().replace(",", "") if price_field else ""
        price_val = int(price_text) if price_text.isdigit() else 0
        yield_rate = 0
        investment = price_val - total_deposit
        if investment > 0: yield_rate = (total_yearly / investment) * 100
        self.lbl_house_totals.setText(f"합계: 보증금 {total_deposit:,}만원 / 월세 {total_monthly:,}만원 (연 {total_yearly:,}만원) 수익률: {yield_rate:.2f}%")

    def generate_marketing_text(self):
        """공개 설정 여부를 반영한 마케팅용 발송 문구 생성"""
        type_name = self.cb_type.currentText()
        m_id = self.le_m_id.text() or "ID 미발급"
        is_public = self.chk_loc_public.isChecked()
        
        addr = f"{self.le_do.text()} {self.le_si_gun.text()} {self.le_emd_ri.text()}"
        detail_loc = ""
        
        if type_name == "아파트":
            complex_name = self.cb_complex.currentText()
            dong = self.cb_dong.currentText()
            ho = self.cb_ho.currentText()
            detail_loc = f"\n📍 위치: {complex_name} {dong}동 " + (f"{ho}호" if is_public else "(호실 비공개)")
        else:
            jibon = self.le_jibon.text()
            detail_loc = f"\n📍 위치: {addr} " + (f"{jibon}" if is_public else "(상세지번 비공개)")

        # 주요 정보 요약 (G2, G3에서 발췌)
        summary = []
        for group in ["G2", "G3"]:
            for k, v in self.all_fields[group].items():
                val = v.currentText() if isinstance(v, QComboBox) else v.text()
                if val and val != "-":
                    summary.append(f"▪ {k}: {val}")

        text = f"""[성균관 부동산 매물안내]
━━━━━━━━━━━━━━
🏠 매물종류: {type_name}
🆔 매물번호: {m_id}{detail_loc}

📋 주요정보
{chr(10).join(summary)}

━━━━━━━━━━━━━━
✨ {type_name} 전문 중개! 
궁금하신 사항은 언제든 문의주세요.\n{Config.SNS_FOOTER}"""
        self.output.clear()
        self.output.append(text)
        QMessageBox.information(self, "문구 생성 완료", "하단 출력창에 발송용 문구가 생성되었습니다. 복사해서 사용하세요.")

    def add_to_multi_list(self):
        """현재 입력된 매물 정보를 요약하여 목록형 문구로 누적 추가"""
        type_name = self.cb_type.currentText()
        m_id = self.le_m_id.text().strip().upper() or "ID미발급"
        addr = self.le_emd_ri.text() or "주소미입력"
        
        # 주요 수치 추출 (면적 및 금액)
        area_w = self.all_fields["G2"].get("면적(㎡)") or self.all_fields["G2"].get("대지면적(㎡)")
        price_w = self.all_fields["G3"].get("매매가") or self.all_fields["G3"].get("희망매매가(만원)")
        
        area = area_w.text() if area_w else "-"
        price = price_w.text() if price_w else "-"
        
        # 토지의 경우 지목 추가
        extra = ""
        if type_name == "토지":
            jm_w = self.all_fields["G2"].get("지목")
            jm = (jm_w.currentText() if isinstance(jm_w, QComboBox) else jm_w.text()) if jm_w else ""
            if jm and jm != "-": extra = f" / {jm}"

        current_text = self.output.toPlainText()
        header = "📢 [성균관 부동산 추천 매물 리스트]\n━━━━━━━━━━━━━━\n"
        footer = f"\n━━━━━━━━━━━━━━\n✨ 문의주시면 친절히 안내해 드립니다.{Config.SNS_FOOTER}"
        
        # 이미 목록이 작성 중인지 확인
        if "추천 매물 리스트" not in current_text:
            new_content = header
        else:
            # 기존 하단 안내 문구(footer) 제거 후 내용만 추출
            new_content = current_text.split("━━━━━━━━━━━━━━\n✨")[0]
            
        item_line = f"📍 {type_name}({m_id}): {addr} | {area}㎡ | {price}만원{extra}\n"
        
        self.output.clear()
        self.output.setText(new_content + item_line + footer)
        self.tabs.setCurrentIndex(0)
        QMessageBox.information(self, "목록 추가", f"{m_id} 매물이 안내 목록에 추가되었습니다.")

    def load_data(self):
        """매물ID를 기반으로 엑셀에서 데이터를 찾아 입력창에 채움"""
        m_id = self.le_m_id.text().strip().upper()
        if not m_id:
            QMessageBox.warning(self, "알림", "조회할 매물ID를 입력해주세요.")
            return

        if not os.path.exists(self.db_path):
            QMessageBox.warning(self, "오류", "데이터베이스 파일이 존재하지 않습니다.")
            return

        try:
            found = False
            with pd.ExcelFile(self.db_path, engine='openpyxl') as reader:
                # 1. 메인 정보 조회 (WHS -> 창고, FCT -> 공장)
                prefix = m_id.split('-')[0] if '-' in m_id else ""
                if prefix != "FCT": return
                logical_name = "공장"
                
                db_sheet = logical_name # 이제 직접 매핑된 시트명 사용
                
                if db_sheet in reader.sheet_names:
                    df = pd.read_excel(reader, sheet_name=db_sheet)
                    row = df[df['매물ID'] == m_id]

                    if not row.empty:
                        found = True
                        data = row.iloc[0].to_dict()
                        
                        # 필드 채우기 시작
                        self.cb_type.setCurrentText(logical_name)
                        self.le_do.setText(str(data.get("도", "")))
                        self.le_si_gun.setText(str(data.get("시/군/구", "")))
                        self.le_emd_ri.setText(str(data.get("읍/면/동/리", "")))
                        self.le_jibon.setText(str(data.get("지번", "")))
                        
                        # 필드 데이터 복원
                        for group_name in ["G2", "G3", "MKT", "G4"]:
                            for label, widget in self.all_fields[group_name].items():
                                if isinstance(widget, QPushButton): continue
                                val = str(data.get(label, ""))
                                if val == "nan": val = ""
                                if isinstance(widget, QComboBox):
                                    widget.setCurrentText(val)
                                else:
                                    widget.setText(val)

                # 2. 공장 상세 테이블 조회
                factory_detail_sheet = f"{db_sheet}상세"
                if factory_detail_sheet in reader.sheet_names:
                    df_detail = pd.read_excel(reader, sheet_name=factory_detail_sheet)
                    rows = df_detail[df_detail['매물ID'] == m_id]
                    self.table_factory.setRowCount(0)
                    headers = ["층", "구조", "용도", "면적(㎡)", "보증금(만원)", "월세(만원)", "비고"]
                    for _, d_row in rows.iterrows():
                        r_idx = self.table_factory.rowCount()
                        self.table_factory.insertRow(r_idx)
                        for c_idx, h in enumerate(headers):
                            val = str(d_row.get(h, ""))
                            if val == "nan": val = ""
                            self.table_factory.setItem(r_idx, c_idx, QTableWidgetItem(val))

            if found:
                self.output.append(f"✅ {m_id} 데이터를 성공적으로 불러왔습니다.")
            else:
                QMessageBox.warning(self, "실패", f"ID '{m_id}'에 해당하는 매물을 찾을 수 없습니다.")
        except Exception as e:
            QMessageBox.critical(self, "오류", f"조회 중 오류 발생: {str(e)}")

    def create_box(self, title, color, labels):
        """글자 겹침 방지를 위해 간격과 여백을 재설계한 함수"""
        box = QGroupBox(title)
        # 1. GroupBox 스타일: margin-top을 40px로 늘려 제목 겹침 방지
        box.setStyleSheet(f"""
            QGroupBox {{ 
                background-color: {color}; 
                border: 1px solid #c0c0c0; 
                border-radius: 8px; 
                margin-top: 20px; 
                font-weight: bold; 
            }}
            QGroupBox::title {{ 
                subcontrol-origin: margin; 
                subcontrol-position: top left;
                left: 10px; 
                padding: 0 10px; 
                top: 0px; 
                font-size: 14px;
                color: #333333;
            }}
        """)
        
        layout = QGridLayout()
        # 행 간격 및 내부 여백 최적화
        layout.setVerticalSpacing(15) # 간격을 12에서 15로 조정하여 가독성 향상
        layout.setHorizontalSpacing(10)
        layout.setContentsMargins(15, 30, 15, 15)
        
        # 드롭다운 데이터 설정
        dropdown_data = {
            "거주형태": ["자가", "임대차"],
            "거래유형": ["매매", "전세", "월세"],
            "성별": ["남", "여"],
            "광고상황": ["광고전", "진행중", "거래완료", "매물철회"],
            "광고사이트": ["한방", "이실장", "부동산써브"],
            "자체광고 채널": ["홈페이지", "블로그", "유튜브"],
            "현재상태": ["거주", "공실"],
            "용도지역": ["제1종전용주거", "제2종전용주거", "제1종일반주거", "제2종일반주거", "제3종일반주거", "준주거", "중심상업", "일반상업", "근린상업", "유통상업", "전용공업지역", "일반공업지역", "준공업지역","보전관리지역", "생산관리지역", "계획관리지역", "농림지역", "자연환경보전지역"],
            "통신사": ["SKT", "KT", "LGU+", "SKT알뜰폰", "KT알뜰폰", "LGU+알뜰폰", "미확인"],
            "연락처구분": ["의뢰인", "임차인"]
        }

        field_dict = {}
        for i, text in enumerate(labels):
            lbl = QLabel(text)
            lbl.setMinimumWidth(90)
            # 라벨이 입력창 중앙에 오도록 수직 정렬
            lbl.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            
            if text == "사진,영상링크":
                edit = QPushButton(f"사진,영상링크 ({len(self.selected_media)})")
                edit.setFixedHeight(32)
                edit.setStyleSheet("background-color: #f39c12; color: white; font-weight: bold; border-radius: 4px;")
                edit.clicked.connect(self.select_media)
                self.btn_add_media = edit
            elif any(key in text for key in dropdown_data):
                # 부분 일치하는 키를 찾아 드롭다운 적용 (예: "용도지역/지구/구역" -> "용도지역")
                match_key = next(key for key in dropdown_data if key in text)
                edit = QComboBox()
                edit.addItems(dropdown_data[match_key])
                edit.setFixedHeight(32)
                edit.setStyleSheet("background-color: white; border: 1px solid #ced4da; border-radius: 4px;")
            else:
                edit = QLineEdit()
                edit.setFixedHeight(32)
                edit.setStyleSheet("background-color: white; border: 1px solid #ced4da; border-radius: 4px; padding: 2px 5px;")
            
            field_dict[text] = edit
            layout.addWidget(lbl, i, 0)
            layout.addWidget(edit, i, 1)
            
        box.setLayout(layout)
        return box, field_dict

    def on_type_changed(self):
        """매물 종류 변경 시 입력 필드를 동적으로 재생성"""
        # 기존 위젯 및 레이아웃 제거
        while self.entry_panel.count():
            item = self.entry_panel.takeAt(0)
            if item.layout():
                while item.layout().count():
                    sub_item = item.layout().takeAt(0)
                    if sub_item.widget():
                        sub_item.widget().deleteLater()
            if item.widget(): item.widget().deleteLater()

        type_name = self.cb_type.currentText()
        for w in self.addr_fields: w.setVisible(True)
        self.factory_detail_group.setVisible(type_name in ["창고", "공장"])

        cfg = self.category_config[type_name]
        self.box2, self.all_fields["G2"] = self.create_box("상세정보", "#f1f8e9", cfg["G2"])
        self.box3, self.all_fields["G3"] = self.create_box("가격조건", "#fff3e0", cfg["G3"])
        self.box_mkt, self.all_fields["MKT"] = self.create_box("마케팅 정보", "#e3f2fd", cfg["MKT"]) # 연파랑색 구역 추가

        g4_labels = ["소유자", "성별", "소유자연락처", "통신사", "연락처구분", "연락처", "비번", "조정가", "중개사메모1", "중개사메모2"]
        self.box4, self.all_fields["G4"] = self.create_box("미공개정보", "#fce4ec", g4_labels)

        # 상세정보(G2)와 가격정보(G3)를 한 열(수직)로 배치
        col1_vbox = QVBoxLayout()
        col1_vbox.addWidget(self.box2)
        col1_vbox.addWidget(self.box3)
        
        self.entry_panel.addLayout(col1_vbox)
        self.entry_panel.addWidget(self.box_mkt)
        self.entry_panel.addWidget(self.box4)

        # 실시간 포맷팅 연결 (콤마 및 하이픈)
        if "거래가" in self.all_fields["G3"]:
            self.all_fields["G3"]["거래가"].textChanged.connect(lambda _: self.format_comma("G3", "거래가"))
            self.all_fields["G3"]["거래가"].textChanged.connect(self.calculate_factory_totals)
        if "보증금/월세" in self.all_fields["G3"]:
            self.all_fields["G3"]["보증금/월세"].textChanged.connect(lambda _: self.format_comma("G3", "보증금/월세"))
        if "희망매매가(만원)" in self.all_fields["G3"]:
            self.all_fields["G3"]["희망매매가(만원)"].textChanged.connect(lambda _: self.format_comma("G3", "희망매매가(만원)"))
            self.all_fields["G3"]["희망매매가(만원)"].textChanged.connect(self.calculate_commercial_totals)
        if "면적(㎡)" in self.all_fields["G2"]:
            self.all_fields["G2"]["면적(㎡)"].textChanged.connect(lambda _: self.format_comma("G2", "면적(㎡)"))
        if "대지면적(㎡)" in self.all_fields["G2"]:
            self.all_fields["G2"]["대지면적(㎡)"].textChanged.connect(lambda _: self.format_comma("G2", "대지면적(㎡)"))
        if "건축면적(㎡)" in self.all_fields["G2"]:
            self.all_fields["G2"]["건축면적(㎡)"].textChanged.connect(lambda _: self.format_comma("G2", "건축면적(㎡)"))
        if "연면적(㎡)" in self.all_fields["G2"]:
            self.all_fields["G2"]["연면적(㎡)"].textChanged.connect(lambda _: self.format_comma("G2", "연면적(㎡)"))
        if "개별공시지가(㎡)" in self.all_fields["G3"]:
            self.all_fields["G3"]["개별공시지가(㎡)"].textChanged.connect(lambda _: self.format_comma("G3", "개별공시지가(㎡)"))
        if "자기자본" in self.all_fields["G3"]:
            self.all_fields["G3"]["자기자본"].textChanged.connect(lambda _: self.format_comma("G3", "자기자본"))
        if "대출금" in self.all_fields["G3"]:
            self.all_fields["G3"]["대출금"].textChanged.connect(lambda _: self.format_comma("G3", "대출금"))
        if "보증금" in self.all_fields["G3"]:
            self.all_fields["G3"]["보증금"].textChanged.connect(lambda _: self.format_comma("G3", "보증금"))
        if "월세" in self.all_fields["G3"]:
            self.all_fields["G3"]["월세"].textChanged.connect(lambda _: self.format_comma("G3", "월세"))
        if "전력(kw/h)" in self.all_fields["G3"]:
            self.all_fields["G3"]["전력(kw/h)"].textChanged.connect(lambda _: self.format_comma("G3", "전력(kw/h)"))
        if "조정가" in self.all_fields["G4"]:
            self.all_fields["G4"]["조정가"].textChanged.connect(lambda _: self.format_comma("G4", "조정가"))
            
        if "소유자연락처" in self.all_fields["G4"]:
            self.all_fields["G4"]["소유자연락처"].textChanged.connect(lambda _: self.format_phone("G4", "소유자연락처"))
        if "연락처" in self.all_fields["G4"]:
            self.all_fields["G4"]["연락처"].textChanged.connect(lambda _: self.format_phone("G4", "연락처"))

    def handle_sns_action(self):
        """SNS 문구 생성 드롭다운 선택에 따른 동작 수행"""
        mode = self.cb_sns_type.currentText()
        if "(개별)" in mode:
            self.generate_marketing_text()
        elif "(다중)" in mode:
            self.add_to_multi_list()

    def format_comma(self, group, field):
        """숫자 입력 시 실시간으로 콤마(,) 추가"""
        edit = self.all_fields[group][field]
        text = edit.text().replace(",", "")
        if text.isdigit():
            edit.blockSignals(True)
            try: edit.setText(format(int(text), ","))
            except (ValueError, TypeError): pass
            edit.blockSignals(False)

    def format_phone(self, group, field_name):
        """전화번호 입력 시 하이픈(-) 자동 삽입"""
        edit = self.all_fields[group][field_name]
        text = re.sub(r'[^0-9]', '', edit.text())
        if len(text) >= 10:
            formatted = f"{text[:3]}-{text[3:7]}-{text[7:]}"
        elif len(text) > 6:
            formatted = f"{text[:3]}-{text[3:6]}-{text[6:]}"
        else: formatted = text
        edit.blockSignals(True)
        edit.setText(formatted)
        edit.blockSignals(False)

    def process_save(self):
        type_name = self.cb_type.currentText()
        
        # 1. ID 결정 (수정 모드인지 신규 모드인지 확인)
        pfx = Config.PREFIX_MAP.get(type_name, "ETC")
        
        m_id = self.le_m_id.text().strip().upper()

        # [ID 변환 로직] 기존 ID가 있고, 접두어가 현재 선택된 종류와 다를 경우 자동 변환
        if m_id:
            current_pfx = m_id.split('-')[0]
            if current_pfx in Config.PREFIX_MAP.values() and current_pfx != pfx:
                old_id = m_id
                parts = m_id.split('-')
                if len(parts) >= 3:
                    m_id = f"{pfx}-{parts[1]}-{parts[2]}"
                    self.output.append(f"🔄 카테고리 변경 감지: ID가 {old_id}에서 {m_id}로 변환되었습니다.")

        if not m_id:
            # 신규 ID 생성 로직
            date_str = datetime.datetime.now().strftime("%y%m%d")
            seq = 1
            if os.path.exists(self.db_path):
                try:
                    with pd.ExcelFile(self.db_path, engine='openpyxl') as reader:
                        max_seq = 0
                        date_pattern = f"-{date_str}-"
                        for s_name in reader.sheet_names:
                            df_tmp = pd.read_excel(reader, sheet_name=s_name)
                            if '매물ID' in df_tmp.columns:
                                matched_ids = df_tmp[df_tmp['매물ID'].astype(str).str.contains(date_pattern)]['매물ID'].tolist()
                                for mid in matched_ids:
                                    try:
                                        num = int(str(mid).split('-')[-1])
                                        if num > max_seq: max_seq = num
                                    except (ValueError, IndexError):
                                        continue
                        seq = max_seq + 1
                except Exception as e:
                    print(f"매물ID 순번 조회 중 오류 (기본값 1 사용): {e}")
            m_id = f"{pfx}-{date_str}-{seq:02d}"

        # 1.5 사진 저장 처리
        media_path_info = "-"
        if self.selected_media:
            # 저장 폴더 생성: property_photos/COM-260525-01/
            target_dir = os.path.join(self.photo_base_dir, m_id)
            if not os.path.exists(target_dir):
                os.makedirs(target_dir)
            
            for idx, src_path in enumerate(self.selected_media):
                ext = os.path.splitext(src_path)[1]
                filename = f"{m_id}_{idx+1}{ext}"
                shutil.copy(src_path, os.path.join(target_dir, filename))
            
            media_path_info = os.path.abspath(target_dir) # 엑셀에는 절대경로 저장

        # 2. 데이터 수집
        record = {"매물ID": m_id, "등록일": datetime.datetime.now().strftime("%Y-%m-%d")}
        is_public = self.chk_loc_public.isChecked()

        # 모든 매물에 대해 소재지 정보 기본 수집
        record.update({"도": self.le_do.text(), "시/군/구": self.le_si_gun.text(), "읍/면/동/리": self.le_emd_ri.text()})

        # 공개 설정일 경우 '호실' 또는 '지번'을 미리 배치 (연녹색 구역)
        if is_public:
            record["지번"] = self.le_jibon.text()
        
        for group_key in ["G2", "G3", "MKT"]:
            group = self.all_fields[group_key]
            for label, widget in group.items():
                if isinstance(widget, QPushButton): continue
                record[label] = widget.currentText() if isinstance(widget, QComboBox) else widget.text()

        record["미디어경로(사진/영상)"] = media_path_info

        # 비공개 설정일 경우 '호실' 또는 '지번'을 여기서 배치 (연분홍색 구역 시작)
        if not is_public:
            record["지번"] = self.le_jibon.text()

        for label, widget in self.all_fields["G4"].items():
            record[label] = widget.currentText() if isinstance(widget, QComboBox) else widget.text()

        # 3. 상세 내역 추출 (주택 추가)
        detail_df = None
        detail_df_factory = None
        factory_details = []
        f_headers = ["층", "구조", "용도", "면적(㎡)", "보증금(만원)", "월세(만원)", "비고"]
        for r in range(self.table_factory.rowCount()):
            row_data = {"매물ID": m_id}
            for c in range(self.table_factory.columnCount()):
                item = self.table_factory.item(r, c)
                row_data[f_headers[c]] = item.text() if item else ""
            factory_details.append(row_data)
        detail_df_factory = pd.DataFrame(factory_details)

        # 3. 엑셀 데이터 누적 저장

        try:
            self.output.append(f"⏳ {m_id} 데이터를 저장 중입니다...")
            QApplication.processEvents() 
            
            # 데이터 수집 순서(입력창 순서)가 반영된 record 기반 DataFrame 생성
            new_df = pd.DataFrame([record])
            sheets = {}

            try:
                if os.path.exists(self.db_path):
                    with pd.ExcelFile(self.db_path, engine='openpyxl') as reader:
                        sheets = {sheet: pd.read_excel(reader, sheet_name=sheet) for sheet in reader.sheet_names}
                    
                    if type_name in sheets:
                        # 기존에 동일 ID가 있으면 삭제 (수정 대응)
                        df_main = sheets[type_name]
                        df_main = df_main[df_main['매물ID'] != m_id]
                        combined = pd.concat([df_main, new_df], ignore_index=True)
                        cols = list(record.keys())
                        for c in combined.columns:
                            if c not in cols: cols.append(c)
                        sheets[type_name] = combined[cols]
                    else:
                        sheets[type_name] = new_df
                    
                    # 상가/주택 상세 DB 처리
                    if detail_df is not None and not detail_df.empty:
                        d_sheet = f"{type_name}상세"
                        if d_sheet in sheets:
                            # 상세 DB도 기존 ID 데이터 삭제 후 추가 (수정 대응)
                            sheets[d_sheet] = sheets[d_sheet][sheets[d_sheet]['매물ID'] != m_id]
                            sheets[d_sheet] = pd.concat([sheets[d_sheet], detail_df], ignore_index=True)
                        else:
                            sheets[d_sheet] = detail_df

                    if detail_df_factory is not None and not detail_df_factory.empty:
                        f_sheet = f"{type_name}상세"
                        if f_sheet in sheets:
                            sheets[f_sheet] = sheets[f_sheet][sheets[f_sheet]['매물ID'] != m_id]
                            sheets[f_sheet] = pd.concat([sheets[f_sheet], detail_df_factory], ignore_index=True)
                        else:
                            sheets[f_sheet] = detail_df_factory

                else:
                    sheets = {type_name: new_df}
                    if detail_df_factory is not None:
                        sheets[f"{type_name}상세"] = detail_df_factory

                # 엑셀 저장 및 서식 적용
                with pd.ExcelWriter(self.db_path, engine='openpyxl') as writer:
                    for s_name, df in sheets.items():
                        df.to_excel(writer, sheet_name=s_name, index=False)
                        ws = writer.sheets[s_name]
                        self._apply_excel_styles(ws, df)
            except PermissionError:
                raise Exception(f"엑셀 파일('{self.db_path}')이 이미 열려 있습니다.\n엑셀 프로그램을 닫고 다시 시도해 주세요.")

            loc_info = f"{record['시/군/구']} {record['읍/면/동/리']}"
            self.output.append(f"✅ {m_id} 저장 성공: {loc_info}")
            QMessageBox.information(self, "성공", f"{m_id} 저장 및 분석이 완료되었습니다.")
            self.clear_inputs() # 저장 성공 후 입력창 초기화
        except Exception as e:
            self.output.append(f"❌ 에러 발생: {str(e)}")
            QMessageBox.critical(self, "저장 에러", f"데이터를 저장하지 못했습니다.\n원인: {str(e)}")

    def _find_private_col(self, df):
        """비공개 정보 시작점 및 특수 컬럼 인덱스 분석"""
        private_col_idx = 999
        media_col_idx = -1
        numeric_cols = []
        for i, col_name in enumerate(df.columns, 1):
            name_str = str(col_name)
            if name_str in ["지번", "소유자"]: private_col_idx = min(private_col_idx, i)
            if name_str == "미디어경로(사진/영상)": media_col_idx = i
            if any(kw in name_str for kw in ["면적", "개별공시지가", "매매가", "보증금", "월세"]):
                numeric_cols.append(i)
        return private_col_idx, media_col_idx, numeric_cols

    def _apply_excel_styles(self, ws, df):
        """엑셀 워크시트에 서식을 적용하는 전용 함수"""
        prv_idx, med_idx, num_cols = self._find_private_col(df)
        style = Config.EXCEL_STYLE
        font_normal = Font(name='맑은 고딕', size=style["font_size"])
        font_bold = Font(name='맑은 고딕', size=style["font_size"], bold=True)
        fill_public = PatternFill(start_color=style["public_color"], end_color=style["public_color"], fill_type="solid")
        fill_private = PatternFill(start_color=style["private_color"], end_color=style["private_color"], fill_type="solid")

        for r_idx, row in enumerate(ws.iter_rows(), 1):
            ws.row_dimensions[r_idx].height = style["row_height"]
            is_header = (r_idx == 1)
            for c_idx, cell in enumerate(row, 1):
                cell.font = font_bold if is_header else font_normal
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.fill = fill_private if c_idx >= prv_idx else fill_public
                
                if not is_header:
                    self._apply_cell_special(cell, c_idx, med_idx, num_cols)
        
        self._auto_column_width(ws)

    def _apply_cell_special(self, cell, col_idx, media_col, numeric_cols):
        """개별 셀 특수 처리 (하이퍼링크 및 숫자 포맷)"""
        if col_idx == media_col and cell.value and cell.value != "-":
            cell.hyperlink = cell.value
            cell.font = Font(name='맑은 고딕', size=Config.EXCEL_STYLE["font_size"], color="0000FF", underline="single")
        elif col_idx in numeric_cols and cell.value and cell.value != "-":
            try:
                cell.value = float(str(cell.value).replace(",", ""))
                cell.number_format = '#,##0'
            except (ValueError, TypeError): pass

    def _auto_column_width(self, ws):
        """내용에 따른 열 너비 자동 조정"""
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    current_len = sum(Config.WIDTH_KOR if ord(c) > 128 else Config.WIDTH_ENG for c in str(cell.value))
                    if current_len > max_len: max_len = current_len
            ws.column_dimensions[col_letter].width = min(max_len + 2, Config.MAX_WIDTH)


if __name__ == '__main__':
    app = QApplication(sys.argv); window = EstateMaster_V13_6(); window.show(); sys.exit(app.exec())