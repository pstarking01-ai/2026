from openpyxl.styles import Font, PatternFill, Alignment
from shared.config import Config


def find_private_col(df, numeric_keywords=None):
    """Identify the private-info start column, the media-path column, and numeric columns."""
    if numeric_keywords is None:
        numeric_keywords = ["면적", "개별공시지가", "매매가", "보증금", "월세"]
    private_col_idx = 999
    media_col_idx = -1
    numeric_cols = []
    for i, col_name in enumerate(df.columns, 1):
        name_str = str(col_name)
        if name_str in ["지번", "소유자"]:
            private_col_idx = min(private_col_idx, i)
        if name_str == "미디어경로(사진/영상)":
            media_col_idx = i
        if any(kw in name_str for kw in numeric_keywords):
            numeric_cols.append(i)
    return private_col_idx, media_col_idx, numeric_cols


def apply_cell_special(cell, col_idx, media_col, numeric_cols):
    """Apply hyperlink or number formatting to individual cells."""
    if col_idx == media_col and cell.value and cell.value != "-":
        cell.hyperlink = cell.value
        cell.font = Font(
            name='맑은 고딕',
            size=Config.EXCEL_STYLE["font_size"],
            color="0000FF",
            underline="single",
        )
    elif col_idx in numeric_cols and cell.value and cell.value != "-":
        try:
            cell.value = float(str(cell.value).replace(",", ""))
            cell.number_format = '#,##0'
        except Exception:
            pass


def apply_excel_styles(ws, df, numeric_keywords=None):
    """Apply consistent formatting to an openpyxl worksheet."""
    prv_idx, med_idx, num_cols = find_private_col(df, numeric_keywords)
    style = Config.EXCEL_STYLE
    font_normal = Font(name='맑은 고딕', size=style["font_size"])
    font_bold = Font(name='맑은 고딕', size=style["font_size"], bold=True)
    fill_public = PatternFill(
        start_color=style["public_color"],
        end_color=style["public_color"],
        fill_type="solid",
    )
    fill_private = PatternFill(
        start_color=style["private_color"],
        end_color=style["private_color"],
        fill_type="solid",
    )

    for r_idx, row in enumerate(ws.iter_rows(), 1):
        ws.row_dimensions[r_idx].height = style["row_height"]
        is_header = (r_idx == 1)
        for c_idx, cell in enumerate(row, 1):
            cell.font = font_bold if is_header else font_normal
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.fill = fill_private if c_idx >= prv_idx else fill_public
            if not is_header:
                apply_cell_special(cell, c_idx, med_idx, num_cols)

    auto_column_width(ws)


def auto_column_width(ws):
    """Auto-adjust column widths based on content length."""
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                current_len = sum(
                    Config.WIDTH_KOR if ord(c) > 128 else Config.WIDTH_ENG
                    for c in str(cell.value)
                )
                if current_len > max_len:
                    max_len = current_len
        ws.column_dimensions[col_letter].width = min(max_len + 2, Config.MAX_WIDTH)
